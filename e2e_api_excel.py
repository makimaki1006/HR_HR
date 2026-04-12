# -*- coding: utf-8 -*-
"""
API直接検証 + Excel出力検証 E2E
- /api/insight/report          (JSON API)
- /api/insight/report/xlsx     (Excel ダウンロード)
- /api/survey/upload           (multipart POST)
- JSONスキーマ安定性（3回連続呼び出し）

認証方式: Playwrightでログイン→Cookie取得→curlで直接API叩く
"""
import os
import sys
import json
import time
import math
import subprocess
import tempfile
from io import BytesIO

BASE = "https://hr-hw.onrender.com"
EMAIL = "test@f-a-c.co.jp"
PASSWORD = "cyxen_2025"
DIR = os.path.dirname(os.path.abspath(__file__))

PASS = 0
FAIL = 0


def check(label, cond, detail=""):
    """1件のチェックをPASS/FAILで記録"""
    global PASS, FAIL
    status = "PASS" if cond else "FAIL"
    icon = "OK" if cond else "NG"
    suffix = f" ({detail})" if detail else ""
    print(f"  [{icon}] [{status}] {label}{suffix}")
    if cond:
        PASS += 1
    else:
        FAIL += 1
    return cond


def curl_get(url, cookie_header, extra_headers=None, binary=False):
    """curlでGET。戻り値: (http_code, headers_dict, body_bytes)"""
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".bin")
    tmp.close()
    hdr_tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".hdr")
    hdr_tmp.close()
    cmd = [
        "curl", "-s", "-o", tmp.name, "-D", hdr_tmp.name,
        "-w", "%{http_code}",
        "-H", f"Cookie: {cookie_header}",
        "-H", "User-Agent: e2e-api-excel/1.0",
    ]
    if extra_headers:
        for k, v in extra_headers.items():
            cmd += ["-H", f"{k}: {v}"]
    cmd.append(url)
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
        code = int(result.stdout.strip() or "0")
        with open(hdr_tmp.name, "r", encoding="utf-8", errors="ignore") as f:
            hdr_raw = f.read()
        headers = {}
        for line in hdr_raw.splitlines():
            if ":" in line:
                k, v = line.split(":", 1)
                headers[k.strip().lower()] = v.strip()
        with open(tmp.name, "rb") as f:
            body = f.read()
        return code, headers, body
    finally:
        try:
            os.unlink(tmp.name)
            os.unlink(hdr_tmp.name)
        except Exception:
            pass


def curl_post_multipart(url, cookie_header, file_path, field_name="csv_file", extra_headers=None):
    """curlでmultipart POST"""
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".bin")
    tmp.close()
    cmd = [
        "curl", "-s", "-o", tmp.name,
        "-w", "%{http_code}",
        "-X", "POST",
        "-H", f"Cookie: {cookie_header}",
        "-H", "User-Agent: e2e-api-excel/1.0",
        "-F", f"{field_name}=@{file_path};type=text/csv",
    ]
    if extra_headers:
        for k, v in extra_headers.items():
            cmd += ["-H", f"{k}: {v}"]
    cmd.append(url)
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
        code = int(result.stdout.strip() or "0")
        with open(tmp.name, "rb") as f:
            body = f.read()
        return code, body
    finally:
        try:
            os.unlink(tmp.name)
        except Exception:
            pass


def has_nan_inf(obj):
    """JSON内にNaN/Infinityが含まれているか再帰チェック"""
    if isinstance(obj, float):
        if math.isnan(obj) or math.isinf(obj):
            return True
    elif isinstance(obj, dict):
        for v in obj.values():
            if has_nan_inf(v):
                return True
    elif isinstance(obj, list):
        for v in obj:
            if has_nan_inf(v):
                return True
    return False


def extract_schema(obj, path=""):
    """再帰的にキーセット（構造シグネチャ）を抽出"""
    if isinstance(obj, dict):
        keys = []
        for k in sorted(obj.keys()):
            keys.append(f"{path}.{k}" if path else k)
            keys.extend(extract_schema(obj[k], f"{path}.{k}" if path else k))
        return keys
    if isinstance(obj, list) and obj:
        # 配列は先頭要素のみでスキーマ化
        return extract_schema(obj[0], f"{path}[]")
    return []


def login_and_get_cookies():
    """Playwrightでログイン → Cookie文字列を返す"""
    from playwright.sync_api import sync_playwright
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        ctx = browser.new_context()
        page = ctx.new_page()
        page.goto(BASE, timeout=60000)
        time.sleep(2)
        page.fill('input[name="email"]', EMAIL)
        page.fill('input[name="password"]', PASSWORD)
        page.click('button[type="submit"]')
        time.sleep(5)
        body = page.text_content("body") or ""
        logged_in = "ログアウト" in body
        cookies = ctx.cookies()
        browser.close()
    if not logged_in:
        print("  [NG] [FAIL] ログインに失敗しました", file=sys.stderr)
        sys.exit(1)
    return "; ".join(f"{c['name']}={c['value']}" for c in cookies)


def create_sample_csv():
    """Indeed形式のサンプルCSV（10行）を作成"""
    path = os.path.join(DIR, "_api_excel_sample.csv")
    rows = [
        "求人タイトル,企業名,勤務地,給与,雇用形態,タグ,URL,新着",
        "営業A,株式会社A,東京都千代田区,月給30万円~35万円,正社員,\"経験者優遇,昇給あり\",https://example.com/1,新着",
        "営業B,株式会社B,東京都千代田区,時給1200円,パート・アルバイト,\"未経験可\",https://example.com/2,新着",
        "営業C,株式会社C,東京都新宿区,月給28万円~32万円,正社員,\"週休2日\",https://example.com/3,新着",
        "営業D,株式会社A,東京都千代田区,月給22万円~27万円,契約社員,\"社保完備\",https://example.com/4,",
        "営業E,株式会社D,東京都渋谷区,時給1300円,パート・アルバイト,\"土日休み\",https://example.com/5,新着",
        "営業F,株式会社B,東京都港区,月給34万円~40万円,正社員,\"昇給あり\",https://example.com/6,",
        "営業G,株式会社E,東京都千代田区,月給25万円~30万円,契約社員,\"未経験可\",https://example.com/7,新着",
        "営業H,株式会社C,東京都新宿区,時給1250円,パート・アルバイト,\"交通費支給\",https://example.com/8,",
        "営業I,株式会社A,東京都渋谷区,月給32万円~38万円,正社員,\"研修制度\",https://example.com/9,新着",
        "営業J,株式会社F,東京都港区,月給27万円~33万円,正社員,\"年間休日120日\",https://example.com/10,",
    ]
    with open(path, "w", encoding="utf-8", newline="") as f:
        f.write("\n".join(rows))
    return path


# ============================================================
# 1. /api/insight/report JSON API
# ============================================================
def test_insight_report_json(cookie_header):
    print("\n=== 1. /api/insight/report JSON API ===")
    code, headers, body = curl_get(f"{BASE}/api/insight/report", cookie_header)

    check("HTTP 200", code == 200, f"actual={code}")
    ct = headers.get("content-type", "")
    check("Content-Type application/json", "application/json" in ct, ct)

    try:
        text = body.decode("utf-8")
        data = json.loads(text)
        check("JSONパース成功", True)
    except Exception as e:
        check("JSONパース成功", False, str(e))
        return None

    # 必須フィールド
    required_fields = [
        "title", "subtitle", "location", "generated_at",
        "executive_summary", "insight_counts", "chapters",
    ]
    missing = [k for k in required_fields if k not in data]
    check("必須フィールド全て存在", not missing, f"missing={missing}")

    # insight_counts サブフィールド
    ic = data.get("insight_counts", {})
    ic_keys = ["critical", "warning", "info", "positive"]
    ic_missing = [k for k in ic_keys if k not in ic]
    check("insight_counts 4種類揃う", not ic_missing, f"missing={ic_missing}")

    # chapters 4章
    chapters = data.get("chapters", [])
    check("chapters.length == 4", len(chapters) == 4, f"len={len(chapters)}")

    # chapters の number が 1,2,3,4 順
    numbers = [c.get("number") for c in chapters]
    check("chapters.number が 1,2,3,4 順", numbers == [1, 2, 3, 4], f"actual={numbers}")

    # insight_counts 合計 == chapters 内の insights 合計
    total_counts = sum(ic.get(k, 0) for k in ic_keys)
    total_chapter_insights = sum(len(c.get("insights", [])) for c in chapters)
    check(
        "insight_counts合計 == chapters内insights合計",
        total_counts == total_chapter_insights,
        f"counts={total_counts} chapters={total_chapter_insights}",
    )

    # NaN/Infinity
    check("NaN/Infinity 非含有", not has_nan_inf(data))

    # Unicode 正常性（日本語タイトル）
    title = data.get("title", "")
    check("Unicode日本語正常", "レポート" in title or "ハローワーク" in title, f"title={title[:30]!r}")

    # chapter insights の構造（代表1件のみ検証）
    sample_insight = None
    for c in chapters:
        if c.get("insights"):
            sample_insight = c["insights"][0]
            break
    if sample_insight is not None:
        has_id = "id" in sample_insight
        has_sev = "severity" in sample_insight
        check("insight.id/severity 存在", has_id and has_sev, f"keys={list(sample_insight.keys())}")
        sev_ok = sample_insight.get("severity") in ("重大", "注意", "情報", "良好")
        check("insight.severity 値が想定ラベル", sev_ok, f"sev={sample_insight.get('severity')!r}")
    else:
        print("  [INFO] insights配列が全章で空のため、個別構造検証はスキップ")

    return data


# ============================================================
# 2. /api/insight/report/xlsx
# ============================================================
def test_insight_report_xlsx(cookie_header):
    print("\n=== 2. /api/insight/report/xlsx ===")
    code, headers, body = curl_get(
        f"{BASE}/api/insight/report/xlsx", cookie_header, binary=True
    )

    check("HTTP 200", code == 200, f"actual={code}")
    ct = headers.get("content-type", "")
    expected_ct = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    check("Content-Type xlsx", expected_ct in ct, ct)

    cd = headers.get("content-disposition", "")
    check("Content-Disposition filename", "hw_report_" in cd and ".xlsx" in cd, cd[:80])

    check("ファイルサイズ > 1KB", len(body) > 1024, f"size={len(body)}B")

    # openpyxl
    try:
        import openpyxl
    except ImportError:
        check("openpyxl 利用可", False, "pip install openpyxl が必要")
        return
    check("openpyxl 利用可", True)

    try:
        wb = openpyxl.load_workbook(BytesIO(body), data_only=False)
        check("Excel 正常に開ける（非破損）", True)
    except Exception as e:
        check("Excel 正常に開ける（非破損）", False, str(e))
        return

    sheet_names = wb.sheetnames
    # 通勤フローは muni が空だと生成されない条件付き。2～3シートのいずれかを許容
    check(
        "シート数 2以上（サマリー/示唆一覧 + 任意で通勤フロー）",
        len(sheet_names) >= 2,
        f"sheets={sheet_names}",
    )
    check("サマリー シート存在", "サマリー" in sheet_names, f"sheets={sheet_names}")
    check("示唆一覧 シート存在", "示唆一覧" in sheet_names, f"sheets={sheet_names}")

    # サマリーシート: タイトル行 + 件数行
    try:
        ws = wb["サマリー"]
        cells_text = []
        for row in ws.iter_rows(max_row=30, values_only=True):
            for v in row:
                if v is not None:
                    cells_text.append(str(v))
        joined = " ".join(cells_text)
        check(
            "サマリー タイトル/件数 記載あり",
            "レポート" in joined or "診断" in joined or "件" in joined,
            f"preview={joined[:60]!r}",
        )
    except Exception as e:
        check("サマリー 読み取り", False, str(e))

    # 示唆一覧: ヘッダー + データ行>=0
    try:
        ws = wb["示唆一覧"]
        rows = list(ws.iter_rows(values_only=True))
        check("示唆一覧 行数 >= 1（ヘッダー）", len(rows) >= 1, f"rows={len(rows)}")
        if rows:
            header = [str(c) if c else "" for c in rows[0]]
            # 期待されるヘッダーの一部を確認
            header_join = " ".join(header)
            expected_keywords = ["ID", "重要度", "タイトル"]
            matched = [k for k in expected_keywords if k in header_join]
            check(
                "示唆一覧 ヘッダー（ID/重要度/タイトル 等）",
                len(matched) >= 2,
                f"header={header} matched={matched}",
            )
            # データ行 >=1 が理想だが、insights 0件の地域もあり得るので警告扱いに留めず情報だけ
            print(f"  [INFO] 示唆一覧 データ行数={len(rows)-1}")
    except Exception as e:
        check("示唆一覧 読み取り", False, str(e))

    # 数式エラー値の検出
    err_tokens = ("#VALUE!", "#NAME?", "#REF!", "#DIV/0!", "#N/A", "#NUM!", "#NULL!")
    found_err = False
    err_location = ""
    for sn in sheet_names:
        ws = wb[sn]
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if isinstance(v, str) and v in err_tokens:
                    found_err = True
                    err_location = f"{sn}: {v}"
                    break
            if found_err:
                break
        if found_err:
            break
    check("数式エラー値 非含有", not found_err, err_location or "")


# ============================================================
# 3. /api/survey/upload 直接POST
# ============================================================
def test_survey_upload(cookie_header):
    print("\n=== 3. /api/survey/upload 直接POST ===")
    csv_path = create_sample_csv()
    try:
        code, body = curl_post_multipart(
            f"{BASE}/api/survey/upload",
            cookie_header,
            csv_path,
            field_name="csv_file",
            extra_headers={"Origin": BASE},
        )

        check("HTTP 200", code == 200, f"actual={code}")

        try:
            html = body.decode("utf-8", errors="replace")
        except Exception as e:
            check("レスポンス decode 可", False, str(e))
            return

        # エラー枠判定
        is_error = "ファイル読み取りエラー" in html or "CSVパースエラー" in html or "CSVファイルが選択" in html
        check("エラー応答ではない", not is_error, html[:100] if is_error else "")

        # 期待HTML要素
        expected_any = ["総求人数", "求人数", "件", "stat-card", "session_id", "s_"]
        matched = [k for k in expected_any if k in html]
        check(
            "レスポンスHTML 分析結果要素含有",
            len(matched) >= 2,
            f"matched={matched}",
        )

        # session_id 抽出
        session_id = None
        import re
        m = re.search(r"(s_[0-9a-f\-]{8,})", html)
        if m:
            session_id = m.group(1)
            print(f"  [INFO] session_id={session_id}")

        # セッションキャッシュ確認: /report/survey?session_id=...
        if session_id:
            code2, _, body2 = curl_get(
                f"{BASE}/report/survey?session_id={session_id}",
                cookie_header,
            )
            check("/report/survey?session_id 200", code2 == 200, f"actual={code2}")
            check(
                "/report/survey キャッシュヒット（HTML>1KB）",
                len(body2) > 1024,
                f"size={len(body2)}",
            )
        else:
            print("  [INFO] session_id抽出できず、キャッシュ検証スキップ")
    finally:
        try:
            os.unlink(csv_path)
        except Exception:
            pass


# ============================================================
# 4. JSONスキーマ安定性
# ============================================================
def test_schema_stability(cookie_header):
    print("\n=== 4. JSONスキーマ安定性（3回連続） ===")
    schemas = []
    for i in range(3):
        code, _, body = curl_get(f"{BASE}/api/insight/report", cookie_header)
        if code != 200:
            check(f"[{i+1}回目] HTTP 200", False, f"code={code}")
            return
        try:
            data = json.loads(body.decode("utf-8"))
        except Exception as e:
            check(f"[{i+1}回目] JSONパース", False, str(e))
            return
        schemas.append(set(extract_schema(data)))
        time.sleep(0.5)

    check("[1回目] スキーマ取得", bool(schemas[0]))
    check("[2回目] スキーマ取得", bool(schemas[1]))
    check("[3回目] スキーマ取得", bool(schemas[2]))

    diff_12 = schemas[0].symmetric_difference(schemas[1])
    diff_13 = schemas[0].symmetric_difference(schemas[2])
    check(
        "1回目と2回目でスキーマ一致",
        not diff_12,
        f"diff={list(diff_12)[:5]}" if diff_12 else "",
    )
    check(
        "1回目と3回目でスキーマ一致",
        not diff_13,
        f"diff={list(diff_13)[:5]}" if diff_13 else "",
    )


# ============================================================
# main
# ============================================================
def main():
    start = time.time()
    print("=" * 50)
    print("API & Excel E2E 検証開始")
    print(f"BASE: {BASE}")
    print("=" * 50)

    print("\n=== 0. ログイン（Cookie取得） ===")
    cookie_header = login_and_get_cookies()
    check("Cookie取得成功", bool(cookie_header), f"len={len(cookie_header)}")

    test_insight_report_json(cookie_header)
    test_insight_report_xlsx(cookie_header)
    test_survey_upload(cookie_header)
    test_schema_stability(cookie_header)

    elapsed = time.time() - start
    print("\n" + "=" * 50)
    print(f"API E2E Summary: PASS: {PASS} / FAIL: {FAIL}  (elapsed={elapsed:.1f}s)")
    print("=" * 50)

    sys.exit(0 if FAIL == 0 else 1)


if __name__ == "__main__":
    main()
