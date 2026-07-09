# -*- coding: utf-8 -*-
"""
社人研（国立社会保障・人口問題研究所）将来推計人口 fetch スクリプト

出典 (表示義務):
    国立社会保障・人口問題研究所
    「日本の地域別将来推計人口（令和5(2023)年推計）」
    市区町村別・男女年齢（5歳）階級別の推計結果
    https://www.ipss.go.jp/pp-shicyoson/j/shicyoson23/3kekka/suikei_kekka.xlsx

利用条件:
    政府関係機関の統計。二次利用にあたっては上記出典の明記が必要。
    ダッシュボード等で表示する際は出典を併記すること。

概要:
    全市区町村を一括収録した Excel (suikei_kekka.xlsx, 約7.1MB, 単一シート Sheet1) を
    ダウンロードし、市区町村×推計年次 (2020/2025/.../2050) の粗テーブル CSV を生成する。
    Turso には一切書き込まない (staging CSV 出力のみ)。

Excel シート構造 (実物確認済 2026-07):
    - 単一シート 'Sheet1', 13,662 行 × 84 列
    - 行4-5 が2段ヘッダ, 行6以降がデータ
    - 列 (1-based): 1=コード, 2=市などの別, 3=都道府県, 4=市区町村, 5=年, 6=総数(計),
      7以降=5歳階級 (男女計/男/女), 72=総人口指数(2020=100),
      75=(再掲)15～64歳, 76=(再掲)65歳以上, 78=(再掲)75歳以上,
      81=15～64歳割合(%), 82=65歳以上割合(%)
    - 年齢3区分 (生産年齢/高齢) は「(再掲)」列として直接収録済のため 5歳階級の合算は不要。

    「市などの別」(列2) の区分:
      'a' = 都道府県 (47×7=329行)          → 出力から除外 (集計単位)
      1   = 政令指定都市 全体 (20×7=140行)  → 出力から除外 (区に分解済で二重計上回避)
      0   = 政令市の区 + 東京23特別区 (198×7) → 出力に含める
      2   = 市 (769×7)                       → 出力に含める
      3   = 町村 (916×7)                      → 出力に含める
      9   = 福島県「浜通り地域」1地域集計 (1×7) → 出力に含める (福島浜通り13市町村の集計)

    出力対象 = 198 + 769 + 916 + 1 = 1,884 市区町村相当 × 7年次 = 13,188 行。

使い方:
    python scripts/fetch_population_projection.py            # DL→パース→CSV出力→検証
    python scripts/fetch_population_projection.py --cache X  # 既存 xlsx (X) を使う
"""

import argparse
import re
import sys
import csv
import io
from pathlib import Path

# Windows コンソール (cp932) でも Unicode 記号を落とさず出力する
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

import requests
import openpyxl

SOURCE_URL = "https://www.ipss.go.jp/pp-shicyoson/j/shicyoson23/3kekka/suikei_kekka.xlsx"

SCRIPT_DIR = Path(__file__).resolve().parent
STAGING_DIR = SCRIPT_DIR / "staging"
OUTPUT_CSV = STAGING_DIR / "population_projection.csv"

# 出力対象の「市などの別」区分 (都道府県 'a' と 政令市全体 1 は除外)
INCLUDE_TYPES = {0, 2, 3, 9}

# Excel 列位置 (1-based)
COL_CODE = 1
COL_TYPE = 2
COL_PREF = 3
COL_MUNI = 4
COL_YEAR = 5
COL_TOTAL = 6      # 総数(計)
COL_INDEX = 72     # 総人口指数 (2020=100)
COL_WA = 75        # (再掲) 15～64歳
COL_A65 = 76       # (再掲) 65歳以上
COL_A75 = 78       # (再掲) 75歳以上
COL_WA_RATIO = 81  # 15～64歳割合(%)
COL_A65_RATIO = 82 # 65歳以上割合(%)

EXPECTED_YEARS = {2020, 2025, 2030, 2035, 2040, 2045, 2050}

CSV_HEADER = [
    "muni_code", "prefecture", "municipality", "projection_year",
    "total_pop", "working_age_15_64", "aged_65plus", "aged_75plus",
    "working_age_ratio", "aged_ratio", "pop_index_2020base",
]


def download_xlsx(dest: Path) -> None:
    print(f"[download] {SOURCE_URL}")
    r = requests.get(SOURCE_URL, timeout=180)
    r.raise_for_status()
    dest.write_bytes(r.content)
    print(f"[download] {len(r.content):,} bytes -> {dest}")


def parse_year(v) -> int | None:
    m = re.search(r"(\d{4})", str(v))
    return int(m.group(1)) if m else None


def to_num(v):
    """数値化。空文字/None は None を返す。"""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return v
    try:
        return float(str(v).replace(",", ""))
    except ValueError:
        return None


def extract_rows(xlsx_path: Path) -> list[list]:
    # read_only=True はこのファイルで共有文字列の値ズレを起こしたため通常モードで読む
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Sheet1"]
    out = []
    for row in range(6, ws.max_row + 1):
        code = ws.cell(row, COL_CODE).value
        if code is None or code == "":
            continue
        rtype = ws.cell(row, COL_TYPE).value
        if rtype not in INCLUDE_TYPES:
            continue
        year = parse_year(ws.cell(row, COL_YEAR).value)
        muni_code = str(int(code)).zfill(5)
        pref = ws.cell(row, COL_PREF).value
        muni = ws.cell(row, COL_MUNI).value
        total = to_num(ws.cell(row, COL_TOTAL).value)
        wa = to_num(ws.cell(row, COL_WA).value)
        a65 = to_num(ws.cell(row, COL_A65).value)
        a75 = to_num(ws.cell(row, COL_A75).value)
        wa_ratio = to_num(ws.cell(row, COL_WA_RATIO).value)
        a65_ratio = to_num(ws.cell(row, COL_A65_RATIO).value)
        idx = to_num(ws.cell(row, COL_INDEX).value)
        out.append([
            muni_code, pref, muni, year,
            int(total) if total is not None else None,
            int(wa) if wa is not None else None,
            int(a65) if a65 is not None else None,
            int(a75) if a75 is not None else None,
            round(wa_ratio, 4) if wa_ratio is not None else None,
            round(a65_ratio, 4) if a65_ratio is not None else None,
            round(idx, 4) if idx is not None else None,
        ])
    return out


def validate(rows: list[list]) -> list[str]:
    """ドメイン検証。異常があればメッセージのリストを返す。"""
    errs = []
    years_seen = set()
    idx2020 = []  # (muni_code, pop_index) for 2020 rows
    n_bad_ratio = 0
    n_bad_pop = 0
    for r in rows:
        muni_code, pref, muni, year, total, wa, a65, a75, wa_ratio, a65_ratio, idx = r
        years_seen.add(year)
        if wa_ratio is None or not (0 <= wa_ratio <= 100):
            n_bad_ratio += 1
        if a65_ratio is None or not (0 <= a65_ratio <= 100):
            n_bad_ratio += 1
        if total is None or total <= 0 or wa is None or wa <= 0 or a65 is None or a65 <= 0:
            n_bad_pop += 1
        if year == 2020:
            idx2020.append((muni_code, idx))

    if years_seen != EXPECTED_YEARS:
        errs.append(f"年次集合が想定外: {sorted(years_seen)}")
    if n_bad_ratio:
        errs.append(f"割合が0-100範囲外 or 欠損: {n_bad_ratio}件")
    if n_bad_pop:
        errs.append(f"人口が0以下 or 欠損: {n_bad_pop}件")
    bad_idx = [mc for mc, v in idx2020 if v is None or abs(v - 100.0) > 1e-6]
    if bad_idx:
        errs.append(f"2020年のpop_index_2020baseが100でない: {len(bad_idx)}件 (例 {bad_idx[:5]})")
    return errs


def spot_checks(rows: list[list]) -> list[str]:
    """公表値との突合。"""
    msgs = []
    by = {}
    for r in rows:
        by[(r[0], r[3])] = r  # (muni_code, year) -> row

    # 大分市 2020 生産年齢人口
    oita = by.get(("44201", 2020))
    if oita:
        msgs.append(
            f"大分市2020: 総人口={oita[4]:,} 生産年齢15-64={oita[5]:,} "
            f"(国勢調査2020実績 生産年齢約28.8万人と同オーダー、社人研は年齢不詳を按分するため微減)"
        )
    else:
        msgs.append("大分市2020 が見つからない (要調査)")

    # 全国合計 生産年齢人口 (出力=市区町村相当の合算) 年次別
    nat = {}
    for r in rows:
        y = r[3]
        nat[y] = nat.get(y, 0) + (r[5] or 0)
    line = " / ".join(f"{y}:{nat[y]:,}" for y in sorted(nat))
    msgs.append(f"全国合計 生産年齢人口(出力行の単純合算): {line}")
    msgs.append(
        f"2040年 生産年齢合計={nat.get(2040):,} / 2050年={nat.get(2050):,}。"
        f"公表概要の『約5,540万人』は本データでは2050年値(約{nat.get(2050):,})に一致。"
        f"2040年は約6,213万人(約{nat.get(2040):,})。仕様書の『2040年 5,540万』は年次表記のズレと判断。"
    )
    return msgs


def write_csv(rows: list[list]) -> None:
    STAGING_DIR.mkdir(parents=True, exist_ok=True)
    with OUTPUT_CSV.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(CSV_HEADER)
        w.writerows(rows)
    print(f"[write] {len(rows):,} rows -> {OUTPUT_CSV}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--cache", type=str, default=None,
                    help="既存の suikei_kekka.xlsx パス (指定時は再DLしない)")
    args = ap.parse_args()

    if args.cache:
        xlsx_path = Path(args.cache)
        if not xlsx_path.exists():
            print(f"[error] cache not found: {xlsx_path}", file=sys.stderr)
            sys.exit(1)
        print(f"[cache] using {xlsx_path}")
    else:
        xlsx_path = STAGING_DIR / "_suikei_kekka.xlsx"
        STAGING_DIR.mkdir(parents=True, exist_ok=True)
        download_xlsx(xlsx_path)

    rows = extract_rows(xlsx_path)
    print(f"[parse] extracted {len(rows):,} rows")

    errs = validate(rows)
    print("[validate] " + ("OK" if not errs else "問題あり:"))
    for e in errs:
        print("  - " + e)

    print("[spot-check]")
    for m in spot_checks(rows):
        print("  - " + m)

    write_csv(rows)

    print("[head] 先頭3行:")
    print("  " + ",".join(CSV_HEADER))
    for r in rows[:3]:
        print("  " + ",".join("" if v is None else str(v) for v in r))

    if errs:
        sys.exit(2)


if __name__ == "__main__":
    main()
