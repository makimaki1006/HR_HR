"""JILPT職業情報DB + 賃金構造基本統計調査 → Turso SQL 出力。

入力（HR_HR/data/jobtag_raw/ に配置済み）:
  - jobtag_desc.csv         JILPT 解説系 ver.7.01
  - jobtag_numeric.csv      JILPT 数値系 ver.7.00
  - table5_age.xlsx         賃金構造基本統計調査 令和7年 表5（職種小分類×年齢階級別）

出力（標準出力 or --output 指定ファイル）:
  Turso `country-statistics` に投入するSQL（CREATE + DELETE + INSERT）

使い方:
  python scripts/import_jobtag_driver.py > data/jobtag_turso_import.sql
  turso db shell country-statistics < data/jobtag_turso_import.sql

設計方針:
  - jobtag_id を主キーとして全職業対応（driver12固定にしない、後から追加可能）
  - category列で職業グループ管理（driver/medical/...）
  - 賃金センサスは wage_census_code 単位で正規化（複数職業が同じcodeを共有）
"""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path
from typing import Any

import openpyxl

HERE = Path(__file__).resolve().parent
DATA_DIR = HERE.parent / "data" / "jobtag_raw"

# ───────────────────────── 対象職業定義 ─────────────────────────

# driver: 道路系/タクシー/バス/トラック/配送/鉄道（後から増やし可、categoryで管理）
DRIVER_OCCUPATIONS: list[dict[str, Any]] = [
    {"jobtag_id": 477, "name": "トラックドライバー",          "wage_census_code": "1614", "category": "driver"},
    {"jobtag_id": 478, "name": "トレーラートラックドライバー",  "wage_census_code": "1614", "category": "driver"},
    {"jobtag_id": 479, "name": "ダンプカー運転手",           "wage_census_code": "1614", "category": "driver"},
    {"jobtag_id": 482, "name": "ルート配送ドライバー",         "wage_census_code": "1703", "category": "driver"},
    {"jobtag_id": 483, "name": "宅配便配達員",              "wage_census_code": "1703", "category": "driver"},
    {"jobtag_id": 532, "name": "フードデリバリー（料理配達員）", "wage_census_code": "1703", "category": "driver"},
    {"jobtag_id": 186, "name": "路線バス運転士",            "wage_census_code": "1611", "category": "driver"},
    {"jobtag_id": 187, "name": "観光バス運転士",            "wage_census_code": "1611", "category": "driver"},
    {"jobtag_id": 480, "name": "送迎バス等運転手",           "wage_census_code": "1611", "category": "driver"},
    {"jobtag_id": 188, "name": "タクシー運転手",            "wage_census_code": "1612", "category": "driver"},
    {"jobtag_id": 481, "name": "介護タクシー運転手",          "wage_census_code": "1612", "category": "driver"},
    {"jobtag_id": 192, "name": "電車運転士",               "wage_census_code": "1601", "category": "driver"},
]

# wage_census_code → 表5 B列の職種名
WAGE_CENSUS_NAME_BY_CODE = {
    "1601": "鉄道運転従事者",
    "1611": "バス運転者",
    "1612": "タクシー運転者",
    "1614": "営業用大型貨物自動車運転者",
    "1703": "その他の運搬従事者",  # ジョブタグAPIで「ルート配送/宅配/フードデリ」が参照
}

# 表5の集計行直下にこの順で12年齢階級が並ぶ
AGE_RANGE_LABELS = [
    "総計", "～19歳", "20～24歳", "25～29歳", "30～34歳", "35～39歳",
    "40～44歳", "45～49歳", "50～54歳", "55～59歳", "60～64歳",
    "65～69歳", "70歳～",
]

# 数値系CSVのカテゴリ→列範囲（行17 = 日本語ラベル）
NUMERIC_CATEGORIES = [
    {"key": "interest", "cols": (4, 9)},
    {"key": "values",   "cols": (10, 20)},
    {"key": "skills",   "cols": (21, 59)},
]

# 解説系CSV: 本文カラム
DESC_TEXT_COLS = {
    "summary":            42,
    "what_is_the_job":    43,
    "how_to_become":      44,
    "working_conditions": 45,
}
DESC_ALIAS_RANGE = (17, 41)         # 別名1-25
DESC_QUALIFICATION_RANGE = (66, 100)  # 関連資格1-35


# ───────────────────────── ヘルパ ─────────────────────────

def _q(text: str | None) -> str:
    """SQLite 文字列リテラルに安全に変換（シングルクォートは''でエスケープ）。"""
    if text is None:
        return "NULL"
    s = str(text).replace("'", "''")
    return f"'{s}'"


def _n(value: Any) -> str:
    """数値またはNULL。空文字や '-' は NULL に。"""
    if value is None:
        return "NULL"
    if isinstance(value, (int, float)):
        return str(value)
    text = str(value).strip()
    if text in {"", "-", "X", "x"}:
        return "NULL"
    try:
        f = float(text)
        return str(f)
    except ValueError:
        return "NULL"


# ───────────────────────── 表5 (賃金センサス) ─────────────────────────

def load_wage_age(xlsx_path: Path) -> dict[str, list[dict[str, Any]]]:
    """wage_census_name → [13行（総計+12階級）]"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    target_names = set(WAGE_CENSUS_NAME_BY_CODE.values())
    result: dict[str, list[dict[str, Any]]] = {}
    for row_idx in range(1, ws.max_row + 1):
        v = ws.cell(row_idx, 2).value
        if not isinstance(v, str) or v.strip() not in target_names:
            continue
        rows = []
        for i in range(13):
            r = row_idx + i
            def cell(c):  # noqa: E306
                return ws.cell(r, c).value
            rows.append({
                "age_range_order": i,
                "age_range": AGE_RANGE_LABELS[i],
                "avg_age": cell(4),
                "tenure_years": cell(5),
                "scheduled_hours": cell(6),
                "overtime_hours": cell(7),
                "monthly_total_thousand_yen": cell(8),
                "monthly_scheduled_thousand_yen": cell(9),
                "annual_bonus_thousand_yen": cell(10),
                "workers_count_tenfold": cell(11),
            })
        result[v.strip()] = rows
    return result


def _annual_salary_man_yen(monthly: Any, bonus: Any) -> float | None:
    try:
        m = float(monthly)
        b = float(bonus)
    except (TypeError, ValueError):
        return None
    return round((m * 12 + b) / 10, 2)


# ───────────────────────── JILPT CSV ─────────────────────────

def _read_csv(path: Path) -> list[list[str]]:
    with open(path, "r", encoding="cp932", errors="replace", newline="") as f:
        return list(csv.reader(f))


def load_descriptions(csv_path: Path, target_ids: set[int]) -> dict[int, dict[str, Any]]:
    rows = _read_csv(csv_path)
    out: dict[int, dict[str, Any]] = {}
    for r in rows[14:]:
        if len(r) < 50:
            continue
        try:
            jid = int((r[2] or "").strip())
        except ValueError:
            continue
        if jid not in target_ids:
            continue
        aliases = [
            (r[i] or "").strip()
            for i in range(DESC_ALIAS_RANGE[0], min(DESC_ALIAS_RANGE[1] + 1, len(r)))
            if (r[i] or "").strip()
        ]
        qualifications = [
            (r[i] or "").strip()
            for i in range(DESC_QUALIFICATION_RANGE[0], min(DESC_QUALIFICATION_RANGE[1] + 1, len(r)))
            if (r[i] or "").strip()
        ]
        description = {
            label: (r[col] or "").strip()
            for label, col in DESC_TEXT_COLS.items()
            if col < len(r)
        }
        out[jid] = {
            "mhlw_classification": (r[4] or "").strip() if len(r) > 4 else "",
            "aliases": aliases,
            "qualifications": qualifications,
            "description": description,
        }
    return out


def load_numeric(csv_path: Path, target_ids: set[int]) -> dict[int, list[dict[str, Any]]]:
    """jobtag_id → [{category, item_order, item, score}]"""
    rows = _read_csv(csv_path)
    label_row = rows[16]
    out: dict[int, list[dict[str, Any]]] = {}
    for r in rows[18:]:
        if len(r) < 60:
            continue
        try:
            jid = int((r[2] or "").strip())
        except ValueError:
            continue
        if jid not in target_ids:
            continue
        items: list[dict[str, Any]] = []
        for cat in NUMERIC_CATEGORIES:
            start, end = cat["cols"]
            order_in_cat = 0
            for col in range(start, end + 1):
                if col >= len(r) or col >= len(label_row):
                    continue
                label = (label_row[col] or "").strip()
                if not label:
                    continue
                v = r[col]
                try:
                    score = float(v) if str(v).strip() not in {"", "-"} else None
                except ValueError:
                    score = None
                if score is None:
                    continue
                order_in_cat += 1
                items.append({
                    "category": cat["key"],
                    "item_order": order_in_cat,
                    "item": label,
                    "score": score,
                })
        out[jid] = items
    return out


# ───────────────────────── SQL 出力 ─────────────────────────

DDL = """\
-- v2_external_jobtag_*: ジョブタグ職業情報DB（JILPT）+ 賃金構造基本統計調査
-- 出典:
--   * 賃金構造基本統計調査 令和7年 一般労働者 職種 表5（厚生労働省、e-Stat 00450091）
--   * 職業情報データベース 解説系 ver.7.01 / 簡易版数値系 ver.7.00（独立行政法人 労働政策研究・研修機構）

CREATE TABLE IF NOT EXISTS v2_external_jobtag_occupation (
    jobtag_id INTEGER PRIMARY KEY,
    name TEXT NOT NULL,
    mhlw_classification TEXT,
    wage_census_code TEXT,
    wage_census_name TEXT,
    category TEXT,
    aliases TEXT,
    updated_at TEXT
);
CREATE INDEX IF NOT EXISTS idx_jobtag_occupation_category ON v2_external_jobtag_occupation(category);
CREATE INDEX IF NOT EXISTS idx_jobtag_occupation_wage_code ON v2_external_jobtag_occupation(wage_census_code);

CREATE TABLE IF NOT EXISTS v2_external_jobtag_description (
    jobtag_id INTEGER PRIMARY KEY,
    summary TEXT,
    what_is_the_job TEXT,
    how_to_become TEXT,
    working_conditions TEXT
);

CREATE TABLE IF NOT EXISTS v2_external_jobtag_scores (
    jobtag_id INTEGER,
    category TEXT,
    item_order INTEGER,
    item TEXT,
    score REAL,
    PRIMARY KEY (jobtag_id, category, item)
);
CREATE INDEX IF NOT EXISTS idx_jobtag_scores_cat ON v2_external_jobtag_scores(jobtag_id, category);

CREATE TABLE IF NOT EXISTS v2_external_jobtag_qualifications (
    jobtag_id INTEGER,
    item_order INTEGER,
    name TEXT,
    PRIMARY KEY (jobtag_id, item_order)
);

CREATE TABLE IF NOT EXISTS v2_external_jobtag_wage_age (
    wage_census_code TEXT,
    wage_census_name TEXT,
    age_range_order INTEGER,
    age_range TEXT,
    avg_age REAL,
    tenure_years REAL,
    scheduled_hours REAL,
    overtime_hours REAL,
    monthly_total_thousand_yen REAL,
    monthly_scheduled_thousand_yen REAL,
    annual_bonus_thousand_yen REAL,
    workers_count_tenfold REAL,
    annual_salary_man_yen REAL,
    survey_year TEXT,
    PRIMARY KEY (wage_census_code, age_range_order)
);
CREATE INDEX IF NOT EXISTS idx_jobtag_wage_age_code ON v2_external_jobtag_wage_age(wage_census_code);
"""


def build_sql(occupations: list[dict[str, Any]], data_dir: Path = DATA_DIR,
              survey_year: str = "令和7年") -> str:
    target_ids = {o["jobtag_id"] for o in occupations}
    target_codes = {o["wage_census_code"] for o in occupations}

    wage_by_name = load_wage_age(data_dir / "table5_age.xlsx")
    descs = load_descriptions(data_dir / "jobtag_desc.csv", target_ids)
    scores = load_numeric(data_dir / "jobtag_numeric.csv", target_ids)

    out: list[str] = [DDL]

    # 既存データの削除（idempotent）
    ids_csv = ",".join(str(i) for i in sorted(target_ids))
    codes_csv = ",".join(_q(c) for c in sorted(target_codes))
    out.append(f"DELETE FROM v2_external_jobtag_occupation WHERE jobtag_id IN ({ids_csv});")
    out.append(f"DELETE FROM v2_external_jobtag_description WHERE jobtag_id IN ({ids_csv});")
    out.append(f"DELETE FROM v2_external_jobtag_scores WHERE jobtag_id IN ({ids_csv});")
    out.append(f"DELETE FROM v2_external_jobtag_qualifications WHERE jobtag_id IN ({ids_csv});")
    out.append(f"DELETE FROM v2_external_jobtag_wage_age WHERE wage_census_code IN ({codes_csv});")
    out.append("")

    # トランザクション
    out.append("BEGIN;")

    # 1. occupation
    for o in occupations:
        jid = o["jobtag_id"]
        code = o["wage_census_code"]
        wage_name = WAGE_CENSUS_NAME_BY_CODE.get(code, "")
        desc = descs.get(jid, {})
        mhlw = desc.get("mhlw_classification", "")
        aliases = "、".join(desc.get("aliases", []))
        out.append(
            "INSERT INTO v2_external_jobtag_occupation "
            "(jobtag_id, name, mhlw_classification, wage_census_code, wage_census_name, category, aliases, updated_at) "
            f"VALUES ({jid}, {_q(o['name'])}, {_q(mhlw)}, {_q(code)}, {_q(wage_name)}, "
            f"{_q(o['category'])}, {_q(aliases)}, datetime('now'));"
        )

    # 2. description
    for o in occupations:
        jid = o["jobtag_id"]
        d = descs.get(jid, {}).get("description", {})
        out.append(
            "INSERT INTO v2_external_jobtag_description "
            "(jobtag_id, summary, what_is_the_job, how_to_become, working_conditions) "
            f"VALUES ({jid}, {_q(d.get('summary'))}, {_q(d.get('what_is_the_job'))}, "
            f"{_q(d.get('how_to_become'))}, {_q(d.get('working_conditions'))});"
        )

    # 3. scores
    for o in occupations:
        jid = o["jobtag_id"]
        for s in scores.get(jid, []):
            out.append(
                "INSERT INTO v2_external_jobtag_scores "
                "(jobtag_id, category, item_order, item, score) "
                f"VALUES ({jid}, {_q(s['category'])}, {s['item_order']}, {_q(s['item'])}, {s['score']});"
            )

    # 4. qualifications
    for o in occupations:
        jid = o["jobtag_id"]
        for i, qname in enumerate(descs.get(jid, {}).get("qualifications", []), start=1):
            out.append(
                "INSERT INTO v2_external_jobtag_qualifications "
                f"(jobtag_id, item_order, name) VALUES ({jid}, {i}, {_q(qname)});"
            )

    # 5. wage_age (codeでユニーク、重複職業がいてもInsert は1回)
    inserted_codes: set[str] = set()
    for o in occupations:
        code = o["wage_census_code"]
        if code in inserted_codes:
            continue
        wage_name = WAGE_CENSUS_NAME_BY_CODE.get(code, "")
        rows = wage_by_name.get(wage_name, [])
        for w in rows:
            annual = _annual_salary_man_yen(
                w["monthly_total_thousand_yen"], w["annual_bonus_thousand_yen"]
            )
            out.append(
                "INSERT INTO v2_external_jobtag_wage_age "
                "(wage_census_code, wage_census_name, age_range_order, age_range, avg_age, tenure_years, "
                "scheduled_hours, overtime_hours, monthly_total_thousand_yen, monthly_scheduled_thousand_yen, "
                "annual_bonus_thousand_yen, workers_count_tenfold, annual_salary_man_yen, survey_year) "
                f"VALUES ({_q(code)}, {_q(wage_name)}, {w['age_range_order']}, {_q(w['age_range'])}, "
                f"{_n(w['avg_age'])}, {_n(w['tenure_years'])}, {_n(w['scheduled_hours'])}, "
                f"{_n(w['overtime_hours'])}, {_n(w['monthly_total_thousand_yen'])}, "
                f"{_n(w['monthly_scheduled_thousand_yen'])}, {_n(w['annual_bonus_thousand_yen'])}, "
                f"{_n(w['workers_count_tenfold'])}, "
                f"{annual if annual is not None else 'NULL'}, {_q(survey_year)});"
            )
        inserted_codes.add(code)

    out.append("COMMIT;")
    out.append("")
    return "\n".join(out)


# ───────────────────────── CLI ─────────────────────────

def main() -> int:
    parser = argparse.ArgumentParser(description="JILPT + 賃金センサス → Turso SQL 出力")
    parser.add_argument("-o", "--output", type=Path, default=None,
                        help="出力ファイル（未指定で標準出力）")
    parser.add_argument("--data-dir", type=Path, default=DATA_DIR,
                        help=f"raw データディレクトリ（デフォルト: {DATA_DIR}）")
    args = parser.parse_args()
    sql = build_sql(DRIVER_OCCUPATIONS, data_dir=args.data_dir)

    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(sql, encoding="utf-8")
        print(f"SQL written: {args.output} ({len(sql):,} chars)", file=sys.stderr)
    else:
        sys.stdout.reconfigure(encoding="utf-8")
        print(sql)
    return 0


if __name__ == "__main__":
    sys.exit(main())
