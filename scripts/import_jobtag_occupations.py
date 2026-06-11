"""JILPT職業情報DB + 賃金構造基本統計調査 → Turso SQL 出力（汎用版）。

入力（HR_HR/data/jobtag_raw/ に配置済み）:
  - jobtag_desc.csv              JILPT 解説系 ver.7.01
  - jobtag_numeric.csv           JILPT 数値系 ver.7.00
  - table5_age.xlsx              賃金構造基本統計調査 令和7年 表5（職種小分類×年齢階級別）
  - candidate_occupations.json   対象職業リスト（IDリスト + category + mhlw_code）
  - wage_census_codes.json       賃金センサスコード対応表（Playwrightで生成）
    形式: {
      "jobtag_id_to_wage_code": {"477": "1614", ...},
      "code_to_table5_name": {"1614": "営業用大型貨物自動車運転者", ...}
    }

出力（--output 指定ファイル or data/jobtag_occupations_turso_import.sql）:
  Turso `country-statistics` に投入するSQL（CREATE + DELETE + INSERT）

使い方:
  python scripts/import_jobtag_occupations.py --category all --output data/jobtag_occupations_turso_import.sql
  turso db shell country-statistics < data/jobtag_occupations_turso_import.sql

設計方針:
  - candidate_occupations.json から動的に職業リストを読み込む
  - driver 12 職業は DRIVER_OCCUPATIONS_LEGACY として固定マッピングを保持し、統合する
  - wage_census_codes.json が存在する場合はそこからコード対応を拡張する
  - category列で職業グループ管理（driver/logistics/manufacturing/construction/cleaning/labor/all）
  - 賃金センサスは wage_census_code 単位で正規化（複数職業が同じcodeを共有）
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path
from typing import Any

import openpyxl

HERE = Path(__file__).resolve().parent
DATA_DIR = HERE.parent / "data" / "jobtag_raw"

# ───────────────────────── driver レガシー固定マッピング ─────────────────────────

# driver 12職業: category=driver のエントリが candidate_occupations.json に無い場合の
# フォールバック定義。candidate_occupations.json の category=driver エントリと統合する。
DRIVER_OCCUPATIONS_LEGACY: list[dict[str, Any]] = [
    {"jobtag_id": 477, "name": "トラックドライバー",           "category": "driver"},
    {"jobtag_id": 478, "name": "トレーラートラックドライバー",   "category": "driver"},
    {"jobtag_id": 479, "name": "ダンプカー運転手",            "category": "driver"},
    {"jobtag_id": 482, "name": "ルート配送ドライバー",          "category": "driver"},
    {"jobtag_id": 483, "name": "宅配便配達員",               "category": "driver"},
    {"jobtag_id": 532, "name": "フードデリバリー（料理配達員）",  "category": "driver"},
    {"jobtag_id": 186, "name": "路線バス運転士",             "category": "driver"},
    {"jobtag_id": 187, "name": "観光バス運転士",             "category": "driver"},
    {"jobtag_id": 480, "name": "送迎バス等運転手",            "category": "driver"},
    {"jobtag_id": 188, "name": "タクシー運転手",             "category": "driver"},
    {"jobtag_id": 481, "name": "介護タクシー運転手",           "category": "driver"},
    {"jobtag_id": 192, "name": "電車運転士",                "category": "driver"},
]

# driver の jobtag_id → wage_census_code 固定マッピング
DRIVER_WAGE_CODE: dict[int, str] = {
    477: "1614", 478: "1614", 479: "1614",
    482: "1703", 483: "1703", 532: "1703",
    186: "1611", 187: "1611", 480: "1611",
    188: "1612", 481: "1612",
    192: "1601",
}

# ───────────────────────── 賃金センサス固定マップ（driver 5コード）─────────────────────────

# 既存 driver 5コードは常に保持する。wage_census_codes.json の code_to_table5_name で
# 上書き・拡張される。
_WAGE_CENSUS_NAME_FIXED: dict[str, str] = {
    "1601": "鉄道運転従事者",
    "1611": "バス運転者",
    "1612": "タクシー運転者",
    "1614": "営業用大型貨物自動車運転者",
    "1703": "その他の運搬従事者",
}

# 表5の集計行直下にこの順で12年齢階級が並ぶ
AGE_RANGE_LABELS = [
    "総計", "～19歳", "20～24歳", "25～29歳", "30～34歳", "35～39歳",
    "40～44歳", "45～49歳", "50～54歳", "55～59歳", "60～64歳",
    "65～69歳", "70歳～",
]

# 数値系CSVのカテゴリ→列範囲（行17 = 日本語ラベル）
NUMERIC_CATEGORIES = [
    {"key": "interest", "cols": (4, 9)},
    {"key": "values",   "cols": (10, 20)},
    {"key": "skills",   "cols": (21, 59)},
]

# 解説系CSV: 本文カラム
DESC_TEXT_COLS = {
    "summary":            42,
    "what_is_the_job":    43,
    "how_to_become":      44,
    "working_conditions": 45,
}
DESC_ALIAS_RANGE = (17, 41)          # 別名1-25
DESC_QUALIFICATION_RANGE = (66, 100)  # 関連資格1-35

# サポートするカテゴリ一覧（CLI --category オプション値）
SUPPORTED_CATEGORIES = ["driver", "logistics", "manufacturing", "construction", "cleaning", "labor", "all"]


# ───────────────────────── ヘルパ ─────────────────────────

def _q(text: str | None) -> str:
    """SQLite 文字列リテラルに安全に変換（シングルクォートは''でエスケープ）。"""
    if text is None:
        return "NULL"
    s = str(text).replace("'", "''")
    return f"'{s}'"


def _n(value: Any) -> str:
    """数値またはNULL。空文字や '-' は NULL に。"""
    if value is None:
        return "NULL"
    if isinstance(value, (int, float)):
        return str(value)
    text = str(value).strip()
    if text in {"", "-", "X", "x"}:
        return "NULL"
    try:
        f = float(text)
        return str(f)
    except ValueError:
        return "NULL"


# ───────────────────────── JSON 動的読込 ─────────────────────────

def load_candidate_occupations(json_path: Path) -> list[dict[str, Any]]:
    """candidate_occupations.json を読み込む。
    各エントリに wage_census_code は含まれていないため、後で付与が必要。
    """
    with open(json_path, "r", encoding="utf-8") as f:
        data: list[dict[str, Any]] = json.load(f)
    return data


def load_wage_census_codes(json_path: Path) -> tuple[dict[int, str], dict[str, str]]:
    """wage_census_codes.json を読み込む。
    ファイルが存在しない場合は空辞書を返す（未生成でもエラーにしない）。

    Returns:
        jobtag_id_to_wage_code: {int(jobtag_id): wage_code}
        code_to_table5_name: {wage_code: table5_name}
    """
    if not json_path.exists():
        return {}, {}
    with open(json_path, "r", encoding="utf-8") as f:
        data: dict[str, Any] = json.load(f)
    raw_id_to_code: dict[str, str] = data.get("jobtag_id_to_wage_code", {})
    id_to_code = {int(k): v for k, v in raw_id_to_code.items()}
    code_to_name: dict[str, str] = data.get("code_to_table5_name", {})
    return id_to_code, code_to_name


def build_wage_census_name_map(
    fixed: dict[str, str],
    code_to_table5_name: dict[str, str],
    target_codes: set[str],
) -> dict[str, str]:
    """WAGE_CENSUS_NAME_BY_CODE を動的に構築する。

    優先度: 固定マップ（driver 5コード） > code_to_table5_name（JSON由来）
    target_codes に含まれ、どちらにも存在しないコードは空文字で補完する。
    Excel 走査は不要（wage_census_codes.json の code_to_table5_name を使用）。
    """
    result: dict[str, str] = dict(fixed)
    for code in target_codes:
        if code not in result:
            result[code] = code_to_table5_name.get(code, "")
    return result


def merge_occupations(
    candidates: list[dict[str, Any]],
    id_to_wage_code: dict[int, str],
    category_filter: str,
) -> list[dict[str, Any]]:
    """candidate_occupations.json の職業と driver レガシーを統合し、
    wage_census_code を付与したリストを返す。

    統合ルール:
    1. DRIVER_OCCUPATIONS_LEGACY の jobtag_id を正規 ID セットとして保持
    2. candidate_occupations.json の category=driver エントリがあれば
       LEGACY のエントリを上書きする（name/mhlw_code の精緻化）
    3. category=driver 以外のエントリについては wage_census_code を
       id_to_wage_code（wage_census_codes.json 由来）から取得する
    4. wage_census_code が取得できない場合は None とし、SQL では NULL になる
    """
    # driver: LEGACY をベースに candidate で上書き
    legacy_by_id: dict[int, dict[str, Any]] = {
        o["jobtag_id"]: dict(o) for o in DRIVER_OCCUPATIONS_LEGACY
    }
    cand_by_id: dict[int, dict[str, Any]] = {
        o["jobtag_id"]: o for o in candidates
    }

    merged: list[dict[str, Any]] = []
    seen_ids: set[int] = set()

    # driver: LEGACY + candidate=driver をマージ
    for jid, legacy in sorted(legacy_by_id.items()):
        entry = dict(legacy)
        if jid in cand_by_id:
            # candidate に存在すれば name/mhlw_code を上書き
            cand = cand_by_id[jid]
            entry["name"] = cand.get("name", legacy["name"])
            entry["mhlw_code"] = cand.get("mhlw_code", "")
        else:
            entry["mhlw_code"] = ""
        entry["wage_census_code"] = DRIVER_WAGE_CODE.get(jid, id_to_wage_code.get(jid))
        merged.append(entry)
        seen_ids.add(jid)

    # candidate に driver category=driver で LEGACY に無い ID があれば追加
    for o in candidates:
        jid = o["jobtag_id"]
        if o.get("category") == "driver" and jid not in seen_ids:
            entry = dict(o)
            entry["wage_census_code"] = DRIVER_WAGE_CODE.get(jid, id_to_wage_code.get(jid))
            merged.append(entry)
            seen_ids.add(jid)

    # driver 以外の candidate を追加
    for o in candidates:
        jid = o["jobtag_id"]
        if o.get("category") == "driver" or jid in seen_ids:
            continue
        entry = dict(o)
        entry["wage_census_code"] = id_to_wage_code.get(jid)
        merged.append(entry)
        seen_ids.add(jid)

    # カテゴリフィルタ適用
    if category_filter != "all":
        merged = [o for o in merged if o.get("category") == category_filter]

    # wage_census_code が None のエントリは空文字に正規化
    for o in merged:
        if o.get("wage_census_code") is None:
            o["wage_census_code"] = ""

    return merged


# ───────────────────────── 表5 (賃金センサス) ─────────────────────────

def load_wage_age(
    xlsx_path: Path,
    wage_census_name_by_code: dict[str, str],
) -> dict[str, list[dict[str, Any]]]:
    """wage_census_name → [13行（総計+12階級）]"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    target_names = set(wage_census_name_by_code.values()) - {""}
    result: dict[str, list[dict[str, Any]]] = {}
    for row_idx in range(1, ws.max_row + 1):
        v = ws.cell(row_idx, 2).value
        if not isinstance(v, str) or v.strip() not in target_names:
            continue
        rows = []
        for i in range(13):
            r = row_idx + i

            def cell(c: int, _r: int = r) -> Any:  # noqa: E306
                return ws.cell(_r, c).value

            rows.append({
                "age_range_order": i,
                "age_range": AGE_RANGE_LABELS[i],
                "avg_age": cell(4),
                "tenure_years": cell(5),
                "scheduled_hours": cell(6),
                "overtime_hours": cell(7),
                "monthly_total_thousand_yen": cell(8),
                "monthly_scheduled_thousand_yen": cell(9),
                "annual_bonus_thousand_yen": cell(10),
                "workers_count_tenfold": cell(11),
            })
        result[v.strip()] = rows
    return result


def _annual_salary_man_yen(monthly: Any, bonus: Any) -> float | None:
    try:
        m = float(monthly)
        b = float(bonus)
    except (TypeError, ValueError):
        return None
    return round((m * 12 + b) / 10, 2)


# ───────────────────────── JILPT CSV ─────────────────────────

def _read_csv(path: Path) -> list[list[str]]:
    with open(path, "r", encoding="cp932", errors="replace", newline="") as f:
        return list(csv.reader(f))


def load_descriptions(csv_path: Path, target_ids: set[int]) -> dict[int, dict[str, Any]]:
    rows = _read_csv(csv_path)
    out: dict[int, dict[str, Any]] = {}
    for r in rows[14:]:
        if len(r) < 50:
            continue
        try:
            jid = int((r[2] or "").strip())
        except ValueError:
            continue
        if jid not in target_ids:
            continue
        aliases = [
            (r[i] or "").strip()
            for i in range(DESC_ALIAS_RANGE[0], min(DESC_ALIAS_RANGE[1] + 1, len(r)))
            if (r[i] or "").strip()
        ]
        qualifications = [
            (r[i] or "").strip()
            for i in range(DESC_QUALIFICATION_RANGE[0], min(DESC_QUALIFICATION_RANGE[1] + 1, len(r)))
            if (r[i] or "").strip()
        ]
        description = {
            label: (r[col] or "").strip()
            for label, col in DESC_TEXT_COLS.items()
            if col < len(r)
        }
        out[jid] = {
            "mhlw_classification": (r[4] or "").strip() if len(r) > 4 else "",
            "aliases": aliases,
            "qualifications": qualifications,
            "description": description,
        }
    return out


def load_numeric(csv_path: Path, target_ids: set[int]) -> dict[int, list[dict[str, Any]]]:
    """jobtag_id → [{category, item_order, item, score}]"""
    rows = _read_csv(csv_path)
    label_row = rows[16]
    out: dict[int, list[dict[str, Any]]] = {}
    for r in rows[18:]:
        if len(r) < 60:
            continue
        try:
            jid = int((r[2] or "").strip())
        except ValueError:
            continue
        if jid not in target_ids:
            continue
        items: list[dict[str, Any]] = []
        for cat in NUMERIC_CATEGORIES:
            start, end = cat["cols"]
            order_in_cat = 0
            for col in range(start, end + 1):
                if col >= len(r) or col >= len(label_row):
                    continue
                label = (label_row[col] or "").strip()
                if not label:
                    continue
                v = r[col]
                try:
                    score = float(v) if str(v).strip() not in {"", "-"} else None
                except ValueError:
                    score = None
                if score is None:
                    continue
                order_in_cat += 1
                items.append({
                    "category": cat["key"],
                    "item_order": order_in_cat,
                    "item": label,
                    "score": score,
                })
        out[jid] = items
    return out


# ───────────────────────── SQL 出力 ─────────────────────────

DDL = """\
-- v2_external_jobtag_*: ジョブタグ職業情報DB（JILPT）+ 賃金構造基本統計調査
-- 出典:
--   * 賃金構造基本統計調査 令和7年 一般労働者 職種 表5（厚生労働省、e-Stat 00450091）
--   * 職業情報データベース 解説系 ver.7.01 / 簡易版数値系 ver.7.00（独立行政法人 労働政策研究・研修機構）

CREATE TABLE IF NOT EXISTS v2_external_jobtag_occupation (
    jobtag_id INTEGER PRIMARY KEY,
    name TEXT NOT NULL,
    mhlw_classification TEXT,
    wage_census_code TEXT,
    wage_census_name TEXT,
    category TEXT,
    aliases TEXT,
    updated_at TEXT
);
CREATE INDEX IF NOT EXISTS idx_jobtag_occupation_category ON v2_external_jobtag_occupation(category);
CREATE INDEX IF NOT EXISTS idx_jobtag_occupation_wage_code ON v2_external_jobtag_occupation(wage_census_code);

CREATE TABLE IF NOT EXISTS v2_external_jobtag_description (
    jobtag_id INTEGER PRIMARY KEY,
    summary TEXT,
    what_is_the_job TEXT,
    how_to_become TEXT,
    working_conditions TEXT
);

CREATE TABLE IF NOT EXISTS v2_external_jobtag_scores (
    jobtag_id INTEGER,
    category TEXT,
    item_order INTEGER,
    item TEXT,
    score REAL,
    PRIMARY KEY (jobtag_id, category, item)
);
CREATE INDEX IF NOT EXISTS idx_jobtag_scores_cat ON v2_external_jobtag_scores(jobtag_id, category);

CREATE TABLE IF NOT EXISTS v2_external_jobtag_qualifications (
    jobtag_id INTEGER,
    item_order INTEGER,
    name TEXT,
    PRIMARY KEY (jobtag_id, item_order)
);

CREATE TABLE IF NOT EXISTS v2_external_jobtag_wage_age (
    wage_census_code TEXT,
    wage_census_name TEXT,
    age_range_order INTEGER,
    age_range TEXT,
    avg_age REAL,
    tenure_years REAL,
    scheduled_hours REAL,
    overtime_hours REAL,
    monthly_total_thousand_yen REAL,
    monthly_scheduled_thousand_yen REAL,
    annual_bonus_thousand_yen REAL,
    workers_count_tenfold REAL,
    annual_salary_man_yen REAL,
    survey_year TEXT,
    PRIMARY KEY (wage_census_code, age_range_order)
);
CREATE INDEX IF NOT EXISTS idx_jobtag_wage_age_code ON v2_external_jobtag_wage_age(wage_census_code);
"""


def build_sql(
    occupations: list[dict[str, Any]],
    wage_census_name_by_code: dict[str, str],
    data_dir: Path = DATA_DIR,
    survey_year: str = "令和7年",
) -> str:
    target_ids = {o["jobtag_id"] for o in occupations}
    target_codes = {o["wage_census_code"] for o in occupations if o.get("wage_census_code")}

    wage_by_name = load_wage_age(data_dir / "table5_age.xlsx", wage_census_name_by_code)
    descs = load_descriptions(data_dir / "jobtag_desc.csv", target_ids)
    scores = load_numeric(data_dir / "jobtag_numeric.csv", target_ids)

    out: list[str] = [DDL]

    # 既存データの削除（idempotent）
    ids_csv = ",".join(str(i) for i in sorted(target_ids))
    codes_csv = ",".join(_q(c) for c in sorted(target_codes)) if target_codes else "''"
    out.append(f"DELETE FROM v2_external_jobtag_occupation WHERE jobtag_id IN ({ids_csv});")
    out.append(f"DELETE FROM v2_external_jobtag_description WHERE jobtag_id IN ({ids_csv});")
    out.append(f"DELETE FROM v2_external_jobtag_scores WHERE jobtag_id IN ({ids_csv});")
    out.append(f"DELETE FROM v2_external_jobtag_qualifications WHERE jobtag_id IN ({ids_csv});")
    out.append(f"DELETE FROM v2_external_jobtag_wage_age WHERE wage_census_code IN ({codes_csv});")
    out.append("")

    # トランザクション
    out.append("BEGIN;")

    # 1. occupation
    for o in occupations:
        jid = o["jobtag_id"]
        code = o.get("wage_census_code") or ""
        wage_name = wage_census_name_by_code.get(code, "") if code else ""
        desc = descs.get(jid, {})
        # mhlw_classification は jobtag_desc.csv 由来を優先、なければ candidate の mhlw_code
        mhlw = desc.get("mhlw_classification") or o.get("mhlw_code", "")
        aliases = "、".join(desc.get("aliases", []))
        out.append(
            "INSERT INTO v2_external_jobtag_occupation "
            "(jobtag_id, name, mhlw_classification, wage_census_code, wage_census_name, category, aliases, updated_at) "
            f"VALUES ({jid}, {_q(o['name'])}, {_q(mhlw)}, {_q(code) if code else 'NULL'}, "
            f"{_q(wage_name) if wage_name else 'NULL'}, "
            f"{_q(o.get('category'))}, {_q(aliases)}, datetime('now'));"
        )

    # 2. description
    for o in occupations:
        jid = o["jobtag_id"]
        d = descs.get(jid, {}).get("description", {})
        out.append(
            "INSERT INTO v2_external_jobtag_description "
            "(jobtag_id, summary, what_is_the_job, how_to_become, working_conditions) "
            f"VALUES ({jid}, {_q(d.get('summary'))}, {_q(d.get('what_is_the_job'))}, "
            f"{_q(d.get('how_to_become'))}, {_q(d.get('working_conditions'))});"
        )

    # 3. scores
    for o in occupations:
        jid = o["jobtag_id"]
        for s in scores.get(jid, []):
            out.append(
                "INSERT INTO v2_external_jobtag_scores "
                "(jobtag_id, category, item_order, item, score) "
                f"VALUES ({jid}, {_q(s['category'])}, {s['item_order']}, {_q(s['item'])}, {s['score']});"
            )

    # 4. qualifications
    for o in occupations:
        jid = o["jobtag_id"]
        for i, qname in enumerate(descs.get(jid, {}).get("qualifications", []), start=1):
            out.append(
                "INSERT INTO v2_external_jobtag_qualifications "
                f"(jobtag_id, item_order, name) VALUES ({jid}, {i}, {_q(qname)});"
            )

    # 5. wage_age (codeでユニーク、重複職業がいてもInsert は1回)
    inserted_codes: set[str] = set()
    for o in occupations:
        code = o.get("wage_census_code") or ""
        if not code or code in inserted_codes:
            continue
        wage_name = wage_census_name_by_code.get(code, "")
        if not wage_name:
            inserted_codes.add(code)
            continue
        rows = wage_by_name.get(wage_name, [])
        for w in rows:
            annual = _annual_salary_man_yen(
                w["monthly_total_thousand_yen"], w["annual_bonus_thousand_yen"]
            )
            out.append(
                "INSERT INTO v2_external_jobtag_wage_age "
                "(wage_census_code, wage_census_name, age_range_order, age_range, avg_age, tenure_years, "
                "scheduled_hours, overtime_hours, monthly_total_thousand_yen, monthly_scheduled_thousand_yen, "
                "annual_bonus_thousand_yen, workers_count_tenfold, annual_salary_man_yen, survey_year) "
                f"VALUES ({_q(code)}, {_q(wage_name)}, {w['age_range_order']}, {_q(w['age_range'])}, "
                f"{_n(w['avg_age'])}, {_n(w['tenure_years'])}, {_n(w['scheduled_hours'])}, "
                f"{_n(w['overtime_hours'])}, {_n(w['monthly_total_thousand_yen'])}, "
                f"{_n(w['monthly_scheduled_thousand_yen'])}, {_n(w['annual_bonus_thousand_yen'])}, "
                f"{_n(w['workers_count_tenfold'])}, "
                f"{annual if annual is not None else 'NULL'}, {_q(survey_year)});"
            )
        inserted_codes.add(code)

    out.append("COMMIT;")
    out.append("")
    return "\n".join(out)


# ───────────────────────── CLI ─────────────────────────

def main() -> int:
    parser = argparse.ArgumentParser(
        description="JILPT + 賃金センサス → Turso SQL 出力（汎用版）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
カテゴリ一覧:
  driver        道路・鉄道系（12職業 LEGACY + candidate統合）
  logistics     物流・操縦・施設管理系
  manufacturing 製造・修理・組立系
  construction  建設・土木系
  cleaning      清掃・廃棄物系
  labor         倉庫・荷役・梱包系
  all           全カテゴリ（デフォルト）

前提ファイル:
  data/jobtag_raw/candidate_occupations.json  （必須）
  data/jobtag_raw/wage_census_codes.json      （任意: Playwrightで生成後に配置）
  data/jobtag_raw/jobtag_desc.csv             （必須）
  data/jobtag_raw/jobtag_numeric.csv          （必須）
  data/jobtag_raw/table5_age.xlsx             （必須）

注意:
  wage_census_codes.json が未生成の場合、category=driver のみ賃金データが付与されます。
  それ以外のカテゴリは wage_census_code が NULL になります。
""",
    )
    parser.add_argument(
        "--category",
        choices=SUPPORTED_CATEGORIES,
        default="all",
        help="出力対象カテゴリ（デフォルト: all）",
    )
    parser.add_argument(
        "--output", "-o",
        type=Path,
        default=Path("data/jobtag_occupations_turso_import.sql"),
        help="SQL出力先ファイル（デフォルト: data/jobtag_occupations_turso_import.sql）",
    )
    parser.add_argument(
        "--data-dir",
        type=Path,
        default=DATA_DIR,
        help=f"raw データディレクトリ（デフォルト: {DATA_DIR}）",
    )
    parser.add_argument(
        "--survey-year",
        default="令和7年",
        help="賃金センサス調査年ラベル（デフォルト: 令和7年）",
    )
    args = parser.parse_args()

    # --- candidate_occupations.json 読み込み ---
    cand_path = args.data_dir / "candidate_occupations.json"
    if not cand_path.exists():
        print(f"エラー: {cand_path} が見つかりません。", file=sys.stderr)
        return 1
    candidates = load_candidate_occupations(cand_path)

    # --- wage_census_codes.json 読み込み（任意）---
    wage_json_path = args.data_dir / "wage_census_codes.json"
    id_to_wage_code, code_to_table5_name = load_wage_census_codes(wage_json_path)
    if not wage_json_path.exists():
        print(
            "警告: wage_census_codes.json が見つかりません。"
            " category=driver のみ賃金データが付与されます。",
            file=sys.stderr,
        )

    # --- 職業リスト統合・フィルタ ---
    occupations = merge_occupations(candidates, id_to_wage_code, args.category)
    if not occupations:
        print(f"警告: category={args.category} に該当する職業が0件です。", file=sys.stderr)

    # --- WAGE_CENSUS_NAME_BY_CODE 動的構築 ---
    target_codes = {o["wage_census_code"] for o in occupations if o.get("wage_census_code")}
    wage_census_name_by_code = build_wage_census_name_map(
        _WAGE_CENSUS_NAME_FIXED, code_to_table5_name, target_codes
    )

    # --- SQL 生成 ---
    sql = build_sql(
        occupations,
        wage_census_name_by_code,
        data_dir=args.data_dir,
        survey_year=args.survey_year,
    )

    # --- 出力 ---
    output_path = args.output
    # 相対パスの場合はスクリプト親ディレクトリ（リポルート）を基準にする
    if not output_path.is_absolute():
        output_path = HERE.parent / output_path
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(sql, encoding="utf-8")
    print(
        f"SQL written: {output_path} ({len(sql):,} chars) "
        f"[category={args.category}, {len(occupations)} occupations]",
        file=sys.stderr,
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
