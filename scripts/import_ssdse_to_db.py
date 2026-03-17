"""
SSDSE-A (市区町村データセット) → hellowork.db インポート
========================================================
統計センターのSSDSE-A CSVから市区町村レベルの外部データを
hellowork.dbにインポートする。

SSDSE-A: https://www.nstac.go.jp/use/literacy/ssdse/
1741市区町村 × 125項目、国勢調査・住民基本台帳等のデータを整備済み

使い方:
    python import_ssdse_to_db.py [--csv PATH] [--db PATH]

デフォルト:
    --csv  scripts/data/update/SSDSE-A-2025.csv
    --db   data/hellowork.db
"""
import sqlite3
import csv
import os
import sys
import argparse

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_CSV = os.path.join(SCRIPT_DIR, "data", "update", "SSDSE-A-2025.csv")
DEFAULT_DB = os.path.join(os.path.dirname(SCRIPT_DIR), "data", "hellowork.db")
DEFAULT_DAYTIME_XLSX = os.path.join(SCRIPT_DIR, "data", "update", "daytime_pop_raw.xlsx")
DEFAULT_CENSUS_AGE_XLSX = os.path.join(SCRIPT_DIR, "data", "census_age_000032142405.xlsx")

# SSDSE-Aのカラムコード → 意味のマッピング
# 出典: SSDSE_ITEM_2025.xlsx
COLUMN_MAP = {
    "A1101":   "total_population",      # 総人口 [人] (国勢調査2020)
    "A110101": "male_population",        # 総人口(男) [人]
    "A110102": "female_population",      # 総人口(女) [人]
    "A1301":   "age_0_14",              # 15歳未満人口 [人]
    "A1302":   "age_15_64",             # 15~64歳人口 [人]
    "A1303":   "age_65_over",           # 65歳以上人口 [人]
    "A130101": "male_0_14",             # 15歳未満人口(男) [人]
    "A130102": "female_0_14",           # 15歳未満人口(女) [人]
    "A130201": "male_15_64",            # 15~64歳人口(男) [人]
    "A130202": "female_15_64",          # 15~64歳人口(女) [人]
    "A130301": "male_65_over",          # 65歳以上人口(男) [人]
    "A130302": "female_65_over",        # 65歳以上人口(女) [人]
    "A141901": "male_75_over",          # 75歳以上人口(男) [人]
    "A141902": "female_75_over",        # 75歳以上人口(女) [人]
    "A1700":   "foreign_population",     # 外国人人口 [人] (国勢調査2020)
    "A5101":   "inflow",                # 転入者数(日本人移動者) [人] (住基2023)
    "A5102":   "outflow",               # 転出者数(日本人移動者) [人] (住基2023)
    # F1108 = 非労働力人口（昼間人口ではない！）→ マッピング対象外
    # 昼間人口データはSSDSE-Aに含まれないため、別途e-Stat国勢調査から取得が必要
}


def safe_int(val):
    """数値変換（カンマ・空文字対応）"""
    if val is None or val == "" or val == "-" or val == "x" or val == "…":
        return 0
    try:
        return int(str(val).replace(",", "").replace(" ", ""))
    except (ValueError, TypeError):
        return 0


def read_ssdse_csv(csv_path):
    """SSDSE-A CSVを読み込み、市区町村別データの辞書リストを返す"""
    # SSDSE-AはShift-JIS (cp932) エンコーディング
    encodings = ["cp932", "utf-8-sig", "utf-8"]

    for enc in encodings:
        try:
            with open(csv_path, encoding=enc) as f:
                reader = csv.reader(f)
                header_codes = next(reader)  # Row 0: column codes
                header_years = next(reader)  # Row 1: reference years

                # カラムインデックスのマッピング構築
                col_idx = {}
                for i, code in enumerate(header_codes):
                    if code in COLUMN_MAP:
                        col_idx[COLUMN_MAP[code]] = i

                # 都道府県・市区町村のインデックス
                # SSDSE-Aの先頭3列: 地域コード, 都道府県, 市区町村
                pref_idx = 1
                muni_idx = 2

                rows = []
                for row in reader:
                    if len(row) < 10:
                        continue
                    pref = row[pref_idx].strip()
                    muni = row[muni_idx].strip()
                    if not pref:
                        continue

                    data = {"prefecture": pref, "municipality": muni}
                    for field, idx in col_idx.items():
                        data[field] = safe_int(row[idx]) if idx < len(row) else 0

                    rows.append(data)

                print(f"  CSV読み込み完了: {len(rows)} 行 (encoding={enc})")

                # 参照年を取得
                ref_years = {}
                for code, field in COLUMN_MAP.items():
                    for i, c in enumerate(header_codes):
                        if c == code and i < len(header_years):
                            ref_years[field] = header_years[i]
                            break

                return rows, ref_years

        except (UnicodeDecodeError, UnicodeError):
            continue

    print(f"ERROR: {csv_path} の読み込みに失敗", file=sys.stderr)
    return None, None


def import_population(db, rows, ref_years):
    """人口データをインポート → v2_external_population"""

    db.execute("DROP TABLE IF EXISTS v2_external_population")
    db.execute("""
        CREATE TABLE v2_external_population (
            prefecture TEXT,
            municipality TEXT,
            total_population INTEGER,
            male_population INTEGER,
            female_population INTEGER,
            age_0_14 INTEGER,
            age_15_64 INTEGER,
            age_65_over INTEGER,
            aging_rate REAL,
            working_age_rate REAL,
            youth_rate REAL,
            reference_date TEXT,
            PRIMARY KEY (prefecture, municipality)
        )
    """)

    ref_date = ref_years.get("total_population", "2020") + "-10-01"

    pop_rows = []
    for r in rows:
        total = r.get("total_population", 0)
        male = r.get("male_population", 0)
        female = r.get("female_population", 0)
        age_0_14 = r.get("age_0_14", 0)
        age_15_64 = r.get("age_15_64", 0)
        age_65 = r.get("age_65_over", 0)

        if total == 0:
            total = age_0_14 + age_15_64 + age_65

        aging_rate = age_65 / total * 100 if total > 0 else None
        working_age_rate = age_15_64 / total * 100 if total > 0 else None
        youth_rate = age_0_14 / total * 100 if total > 0 else None

        pop_rows.append((
            r["prefecture"], r["municipality"],
            total, male, female,
            age_0_14, age_15_64, age_65,
            aging_rate, working_age_rate, youth_rate,
            ref_date,
        ))

    db.executemany(
        "INSERT OR REPLACE INTO v2_external_population VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
        pop_rows,
    )
    db.execute("CREATE INDEX IF NOT EXISTS idx_ext_pop_pref ON v2_external_population(prefecture)")
    print(f"    v2_external_population: {len(pop_rows)} 行")
    return pop_rows


def import_migration(db, rows, ref_years, pop_rows):
    """社会動態データをインポート → v2_external_migration"""

    db.execute("DROP TABLE IF EXISTS v2_external_migration")
    db.execute("""
        CREATE TABLE v2_external_migration (
            prefecture TEXT,
            municipality TEXT,
            inflow INTEGER,
            outflow INTEGER,
            net_migration INTEGER,
            net_migration_rate REAL,
            reference_year INTEGER,
            PRIMARY KEY (prefecture, municipality)
        )
    """)

    ref_year = int(ref_years.get("inflow", "2023"))

    # 人口マップ（転入転出率の分母）
    pop_map = {}
    for pr in pop_rows:
        pop_map[(pr[0], pr[1])] = pr[2]  # (pref, muni) -> total_pop

    mig_rows = []
    for r in rows:
        inflow = r.get("inflow", 0)
        outflow = r.get("outflow", 0)

        if inflow == 0 and outflow == 0:
            continue

        net = inflow - outflow
        pop = pop_map.get((r["prefecture"], r["municipality"]), 0)
        rate = net / pop * 1000 if pop > 0 else None  # permille

        mig_rows.append((
            r["prefecture"], r["municipality"],
            inflow, outflow, net, rate, ref_year,
        ))

    db.executemany(
        "INSERT OR REPLACE INTO v2_external_migration VALUES (?,?,?,?,?,?,?)",
        mig_rows,
    )
    db.execute("CREATE INDEX IF NOT EXISTS idx_ext_mig_pref ON v2_external_migration(prefecture)")
    print(f"    v2_external_migration: {len(mig_rows)} 行")


def import_foreign_residents(db, rows, ref_years, pop_rows):
    """外国人住民データをインポート → v2_external_foreign_residents"""

    db.execute("DROP TABLE IF EXISTS v2_external_foreign_residents")
    db.execute("""
        CREATE TABLE v2_external_foreign_residents (
            prefecture TEXT,
            municipality TEXT,
            total_foreign INTEGER,
            foreign_rate REAL,
            reference_date TEXT,
            PRIMARY KEY (prefecture, municipality)
        )
    """)

    ref_date = ref_years.get("foreign_population", "2020") + "-10-01"

    pop_map = {}
    for pr in pop_rows:
        pop_map[(pr[0], pr[1])] = pr[2]

    foreign_rows = []
    for r in rows:
        total_foreign = r.get("foreign_population", 0)
        pop = pop_map.get((r["prefecture"], r["municipality"]), 0)
        rate = total_foreign / pop * 100 if pop > 0 else None

        foreign_rows.append((
            r["prefecture"], r["municipality"],
            total_foreign, rate, ref_date,
        ))

    db.executemany(
        "INSERT OR REPLACE INTO v2_external_foreign_residents VALUES (?,?,?,?,?)",
        foreign_rows,
    )
    db.execute("CREATE INDEX IF NOT EXISTS idx_ext_foreign_pref ON v2_external_foreign_residents(prefecture)")
    print(f"    v2_external_foreign_residents: {len(foreign_rows)} 行")


def parse_census_age_excel(xlsx_path):
    """国勢調査Excelから都道府県別×男女×各歳人口を読み込み、10歳階級に集約

    Returns:
        dict[(pref_code, gender), dict[age_group, population]]
        gender: 'male' or 'female'
        age_group: '0-9', '10-19', ..., '70-79', '80+'
    """
    try:
        import openpyxl
    except ImportError:
        print("    国勢調査Excel解析: スキップ (openpyxlが必要)")
        return None

    if not os.path.exists(xlsx_path):
        print(f"    国勢調査Excel: ファイル未配置 ({xlsx_path})")
        return None

    print(f"  国勢調査Excel読み込み: {os.path.basename(xlsx_path)}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sh = wb[wb.sheetnames[0]]

    # 都道府県コード→都道府県名マッピング
    PREF_MAP = {
        "01000": "北海道", "02000": "青森県", "03000": "岩手県", "04000": "宮城県",
        "05000": "秋田県", "06000": "山形県", "07000": "福島県", "08000": "茨城県",
        "09000": "栃木県", "10000": "群馬県", "11000": "埼玉県", "12000": "千葉県",
        "13000": "東京都", "14000": "神奈川県", "15000": "新潟県", "16000": "富山県",
        "17000": "石川県", "18000": "福井県", "19000": "山梨県", "20000": "長野県",
        "21000": "岐阜県", "22000": "静岡県", "23000": "愛知県", "24000": "三重県",
        "25000": "滋賀県", "26000": "京都府", "27000": "大阪府", "28000": "兵庫県",
        "29000": "奈良県", "30000": "和歌山県", "31000": "鳥取県", "32000": "島根県",
        "33000": "岡山県", "34000": "広島県", "35000": "山口県", "36000": "徳島県",
        "37000": "香川県", "38000": "愛媛県", "39000": "高知県", "40000": "福岡県",
        "41000": "佐賀県", "42000": "長崎県", "43000": "熊本県", "44000": "大分県",
        "45000": "宮崎県", "46000": "鹿児島県", "47000": "沖縄県",
    }

    # 10歳階級の定義: (age_group, start_age, end_age)
    AGE_GROUPS = [
        ("0-9", 0, 9), ("10-19", 10, 19), ("20-29", 20, 29),
        ("30-39", 30, 39), ("40-49", 40, 49), ("50-59", 50, 59),
        ("60-69", 60, 69), ("70-79", 70, 79), ("80+", 80, 999),
    ]

    # 各歳データはCol 5(0歳)〜Col 115(110歳以上)
    # Col 5+age = age歳の人口 (0〜109歳)
    # Col 115 = 110歳以上
    AGE_COL_START = 5  # 0歳のカラム
    AGE_COL_110PLUS = 115  # 110歳以上

    result = {}  # (pref_name, gender) -> {age_group: count}

    for r in range(12, sh.max_row + 1):
        c1 = str(sh.cell(r, 1).value or "")
        c2 = str(sh.cell(r, 2).value or "")
        c3 = str(sh.cell(r, 3).value or "")

        # 国籍総数のみ（col1が0_で始まる）
        if not c1.startswith("0_"):
            continue

        # 男女判定
        if c2.startswith("1_"):
            gender = "male"
        elif c2.startswith("2_"):
            gender = "female"
        else:
            continue

        # 都道府県コード抽出（全国行は除外）
        pref_code = c3.split("_")[0] if "_" in c3 else ""
        if pref_code == "00000" or pref_code not in PREF_MAP:
            continue

        pref_name = PREF_MAP[pref_code]

        # 各歳データを読み込み、10歳階級に集約
        age_counts = {}
        for grp_name, start, end in AGE_GROUPS:
            total = 0
            for age in range(start, min(end + 1, 110)):
                col = AGE_COL_START + age
                total += safe_int(sh.cell(r, col).value)
            # 80+グループには110歳以上も加算
            if grp_name == "80+":
                total += safe_int(sh.cell(r, AGE_COL_110PLUS).value)
            age_counts[grp_name] = total

        result[(pref_name, gender)] = age_counts

    wb.close()

    pref_count = len(set(k[0] for k in result.keys()))
    print(f"    国勢調査: {pref_count} 都道府県 × 男女 × 9区分 読み込み完了")
    return result


def import_population_pyramid(db, rows, census_ratios=None):
    """人口ピラミッド（9区分×男女）をインポート → v2_external_population_pyramid

    国勢調査の都道府県別年齢分布比率で、SSDSE-Aの市区町村4区分を9区分に按分推定。
    census_ratiosがNoneの場合は旧4区分にフォールバック。
    """

    db.execute("DROP TABLE IF EXISTS v2_external_population_pyramid")
    db.execute("""
        CREATE TABLE v2_external_population_pyramid (
            prefecture TEXT,
            municipality TEXT,
            age_group TEXT,
            male_count INTEGER,
            female_count INTEGER,
            PRIMARY KEY (prefecture, municipality, age_group)
        )
    """)

    AGE_GROUPS_9 = ["0-9", "10-19", "20-29", "30-39", "40-49",
                    "50-59", "60-69", "70-79", "80+"]

    pyramid_rows = []
    for r in rows:
        pref = r["prefecture"]
        muni = r["municipality"]

        male_0_14 = r.get("male_0_14", 0)
        female_0_14 = r.get("female_0_14", 0)
        male_15_64 = r.get("male_15_64", 0)
        female_15_64 = r.get("female_15_64", 0)
        male_65 = r.get("male_65_over", 0)
        female_65 = r.get("female_65_over", 0)
        male_75 = r.get("male_75_over", 0)
        female_75 = r.get("female_75_over", 0)

        male_65_74 = max(0, male_65 - male_75)
        female_65_74 = max(0, female_65 - female_75)

        total = male_0_14 + female_0_14 + male_15_64 + female_15_64 + male_65 + female_65
        if total == 0:
            continue

        if census_ratios is None:
            # フォールバック: 旧4区分
            pyramid_rows.extend([
                (pref, muni, "0-14", male_0_14, female_0_14),
                (pref, muni, "15-64", male_15_64, female_15_64),
                (pref, muni, "65-74", male_65_74, female_65_74),
                (pref, muni, "75+", male_75, female_75),
            ])
            continue

        # 国勢調査比率で按分
        for gender, ssdse_brackets, census_key in [
            ("male",
             {"0-14": male_0_14, "15-64": male_15_64,
              "65-74": male_65_74, "75+": male_75},
             "male"),
            ("female",
             {"0-14": female_0_14, "15-64": female_15_64,
              "65-74": female_65_74, "75+": female_75},
             "female"),
        ]:
            census = census_ratios.get((pref, census_key))
            if census is None:
                # 都道府県の国勢調査データがない場合、均等按分
                census = {g: 1 for g in AGE_GROUPS_9}

            # 国勢調査の各ブラケット内合計を計算
            c_0_14 = census.get("0-9", 0) + census.get("10-19", 0) * 0.5  # 10-14のみ
            c_0_14_full = sum(census.get(g, 0) for g in ["0-9"]) + census.get("10-19", 0) * 0.5
            # 正確に: 0-14 = 0-9 全部 + 10-19のうち10-14 ≒ 10-19の半分
            # 国勢調査の各歳データから正確な区分を使う
            c_0_9 = census.get("0-9", 0)
            c_10_19 = census.get("10-19", 0)
            c_20_29 = census.get("20-29", 0)
            c_30_39 = census.get("30-39", 0)
            c_40_49 = census.get("40-49", 0)
            c_50_59 = census.get("50-59", 0)
            c_60_69 = census.get("60-69", 0)
            c_70_79 = census.get("70-79", 0)
            c_80plus = census.get("80+", 0)

            # SSDSE-Aの4区分に対応する国勢調査合計
            # 0-14歳 → census 0-9 + census 10-14 (= 10-19の半分と近似)
            # ただし各歳データを持っているので正確に計算可能
            # parse_census_age_excelが10歳階級に集約済みなので、
            # 15-64に対応する国勢調査合計 = 10-19の半分 + 20-29 + ... + 60-69の半分
            # しかし10歳階級境界がSSDE-Aの4区分境界(0-14, 15-64, 65-74, 75+)と
            # きれいに一致しないので、各10歳階級をSSDE-A区分にマッピングする比率を使う

            # 各10歳階級がどのSSDE-A区分に属するかの比率
            # 10-19歳: 10-14は「0-14」、15-19は「15-64」 → 半々と近似
            # 60-69歳: 60-64は「15-64」、65-69は「65-74」 → 半々と近似
            # 70-79歳: 70-74は「65-74」、75-79は「75+」 → 半々と近似

            # SSDSE-A区分ごとの国勢調査合計（按分分母）
            census_in_0_14 = c_0_9 + c_10_19 * 0.5
            census_in_15_64 = c_10_19 * 0.5 + c_20_29 + c_30_39 + c_40_49 + c_50_59 + c_60_69 * 0.5
            census_in_65_74 = c_60_69 * 0.5 + c_70_79 * 0.5
            census_in_75plus = c_70_79 * 0.5 + c_80plus

            # 各10歳階級の推定人口を計算
            estimated = {}

            # 0-9: 0-14区分から按分
            if census_in_0_14 > 0:
                estimated["0-9"] = ssdse_brackets["0-14"] * c_0_9 / census_in_0_14
            else:
                estimated["0-9"] = 0

            # 10-19: 0-14区分の残り + 15-64区分の一部
            ratio_10_14_in_0_14 = (c_10_19 * 0.5) / census_in_0_14 if census_in_0_14 > 0 else 0
            ratio_15_19_in_15_64 = (c_10_19 * 0.5) / census_in_15_64 if census_in_15_64 > 0 else 0
            estimated["10-19"] = (ssdse_brackets["0-14"] * ratio_10_14_in_0_14
                                  + ssdse_brackets["15-64"] * ratio_15_19_in_15_64)

            # 20-29, 30-39, 40-49, 50-59: 15-64区分内で按分
            for grp, c_val in [("20-29", c_20_29), ("30-39", c_30_39),
                                ("40-49", c_40_49), ("50-59", c_50_59)]:
                if census_in_15_64 > 0:
                    estimated[grp] = ssdse_brackets["15-64"] * c_val / census_in_15_64
                else:
                    estimated[grp] = 0

            # 60-69: 15-64の残り + 65-74の一部
            ratio_60_64_in_15_64 = (c_60_69 * 0.5) / census_in_15_64 if census_in_15_64 > 0 else 0
            ratio_65_69_in_65_74 = (c_60_69 * 0.5) / census_in_65_74 if census_in_65_74 > 0 else 0
            estimated["60-69"] = (ssdse_brackets["15-64"] * ratio_60_64_in_15_64
                                  + ssdse_brackets["65-74"] * ratio_65_69_in_65_74)

            # 70-79: 65-74の残り + 75+の一部
            ratio_70_74_in_65_74 = (c_70_79 * 0.5) / census_in_65_74 if census_in_65_74 > 0 else 0
            ratio_75_79_in_75plus = (c_70_79 * 0.5) / census_in_75plus if census_in_75plus > 0 else 0
            estimated["70-79"] = (ssdse_brackets["65-74"] * ratio_70_74_in_65_74
                                  + ssdse_brackets["75+"] * ratio_75_79_in_75plus)

            # 80+: 75+の残り
            if census_in_75plus > 0:
                estimated["80+"] = ssdse_brackets["75+"] * c_80plus / census_in_75plus
            else:
                estimated["80+"] = 0

            # 整数化（四捨五入）
            int_values = {g: round(estimated[g]) for g in AGE_GROUPS_9}

            # 端数調整: 合計をSSDE-Aの男女合計に一致させる
            ssdse_total = (ssdse_brackets["0-14"] + ssdse_brackets["15-64"]
                           + ssdse_brackets["65-74"] + ssdse_brackets["75+"])
            diff = ssdse_total - sum(int_values.values())
            if diff != 0:
                # 最大グループに端数を加算
                max_grp = max(int_values, key=int_values.get)
                int_values[max_grp] += diff

            # pyramid_rowsに追加（genderごとにmale/femaleを振り分け）
            if gender == "male":
                for g in AGE_GROUPS_9:
                    # (pref, muni, age_group, male, female=placeholder)
                    pyramid_rows.append((pref, muni, g, int_values[g], None))
            else:
                # femaleを既存のmale行に結合
                # male行は直前に追加されているので、末尾9行を更新
                base = len(pyramid_rows) - 9
                for i, g in enumerate(AGE_GROUPS_9):
                    old = pyramid_rows[base + i]
                    pyramid_rows[base + i] = (old[0], old[1], old[2], old[3], int_values[g])

    db.executemany(
        "INSERT OR REPLACE INTO v2_external_population_pyramid VALUES (?,?,?,?,?)",
        pyramid_rows,
    )
    db.execute("CREATE INDEX IF NOT EXISTS idx_ext_pyramid_pref ON v2_external_population_pyramid(prefecture)")
    n_groups = 9 if census_ratios else 4
    print(f"    v2_external_population_pyramid: {len(pyramid_rows)} 行 ({len(pyramid_rows)//n_groups} 市区町村 × {n_groups}区分)")


def build_code_to_name_map(csv_path):
    """SSDSE-AのCSVから地域コード→(都道府県, 市区町村)のマッピングを構築"""
    encodings = ["cp932", "utf-8-sig", "utf-8"]
    for enc in encodings:
        try:
            code_map = {}
            with open(csv_path, encoding=enc) as f:
                reader = csv.reader(f)
                next(reader)  # header codes
                next(reader)  # years
                next(reader)  # column names (Japanese)
                for row in reader:
                    if len(row) < 3:
                        continue
                    code = row[0]  # e.g. R01100
                    pref = row[1].strip()
                    muni = row[2].strip()
                    if code.startswith("R") and pref:
                        numeric_code = code[1:]  # 01100
                        code_map[numeric_code] = (pref, muni)
            return code_map
        except (UnicodeDecodeError, UnicodeError):
            continue
    return {}


def import_daytime_population(db, xlsx_path, code_map):
    """国勢調査 昼夜間人口Excelからインポート → v2_external_daytime_population

    データソース: e-Stat 国勢調査 従業地・通学地集計 令和2年(2020)
    statInfId: 000032217063 (不詳補完値)
    """
    try:
        import openpyxl
    except ImportError:
        print("    v2_external_daytime_population: スキップ (openpyxlが必要)")
        return

    if not os.path.exists(xlsx_path):
        # .xls でも試行（実体がxlsx形式の場合がある）
        xls_path = xlsx_path.replace(".xlsx", ".xls")
        if os.path.exists(xls_path):
            xlsx_path = xls_path
        else:
            print(f"    v2_external_daytime_population: スキップ (ファイル未配置: {xlsx_path})")
            return

    print(f"  昼夜間人口Excel読み込み: {os.path.basename(xlsx_path)}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # 令和2年 不詳補完値 シートを探す（最後のシート）
    sh = wb[wb.sheetnames[-1]]

    # データ読み込み: Col3=都道府県コード, Col4=市区町村コード,
    # Col5=夜間人口(常住), Col14=昼間人口(従業地・通学地)
    excel_data = {}
    for r in range(11, sh.max_row + 1):
        v4 = sh.cell(r, 4).value
        if not v4 or "_" not in str(v4):
            continue

        muni_code = str(v4).split("_")[0]
        pref_code = str(sh.cell(r, 3).value).split("_")[0] if sh.cell(r, 3).value else ""

        # 都道府県合計行はスキップ（コードが同じ場合）
        if muni_code == pref_code:
            continue

        nighttime = sh.cell(r, 5).value
        daytime = sh.cell(r, 14).value

        if nighttime is None or daytime is None:
            continue

        # 秘匿値「-」「x」「…」等への対応
        try:
            nighttime = int(nighttime)
            daytime = int(daytime)
        except (ValueError, TypeError):
            continue

        if nighttime == 0 and daytime == 0:
            continue

        excel_data[muni_code] = (nighttime, daytime)

    wb.close()
    print(f"    Excel: {len(excel_data)} 市区町村の昼夜間人口を読み込み")

    # DBテーブル作成
    db.execute("DROP TABLE IF EXISTS v2_external_daytime_population")
    db.execute("""
        CREATE TABLE v2_external_daytime_population (
            prefecture TEXT,
            municipality TEXT,
            nighttime_pop INTEGER,
            daytime_pop INTEGER,
            day_night_ratio REAL,
            inflow_pop INTEGER,
            outflow_pop INTEGER,
            reference_year INTEGER,
            PRIMARY KEY (prefecture, municipality)
        )
    """)

    daytime_rows = []
    matched = 0
    for muni_code, (nighttime, daytime) in excel_data.items():
        if muni_code not in code_map:
            continue

        pref, muni = code_map[muni_code]
        matched += 1

        ratio = daytime / nighttime * 100 if nighttime > 0 else None
        inflow = max(0, daytime - nighttime) if daytime > nighttime else 0
        outflow = max(0, nighttime - daytime) if nighttime > daytime else 0

        daytime_rows.append((
            pref, muni,
            nighttime, daytime, ratio, inflow, outflow, 2020,
        ))

    db.executemany(
        "INSERT OR REPLACE INTO v2_external_daytime_population VALUES (?,?,?,?,?,?,?,?)",
        daytime_rows,
    )
    db.execute("CREATE INDEX IF NOT EXISTS idx_ext_daytime_pref ON v2_external_daytime_population(prefecture)")
    print(f"    v2_external_daytime_population: {len(daytime_rows)} 行 (マッチ: {matched}/{len(excel_data)})")


def verify(db):
    """インポート検証"""
    print("\n=== 検証 ===")
    tables = [
        "v2_external_population",
        "v2_external_migration",
        "v2_external_foreign_residents",
        "v2_external_daytime_population",
        "v2_external_population_pyramid",
    ]
    for table in tables:
        try:
            count = db.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
            prefs = db.execute(f"SELECT COUNT(DISTINCT prefecture) FROM {table}").fetchone()[0]
            print(f"  {table}: {count} 行, {prefs} 都道府県")
        except Exception:
            print(f"  {table}: 未作成")

    # 東京都サンプル
    try:
        row = db.execute("""
            SELECT total_population, aging_rate, working_age_rate, youth_rate
            FROM v2_external_population WHERE prefecture = '東京都' AND municipality = '特別区部'
        """).fetchone()
        if row:
            print(f"\n  東京都特別区部: 人口{row[0]:,}人, 高齢化率{row[1]:.1f}%, 生産年齢{row[2]:.1f}%, 年少{row[3]:.1f}%")
    except Exception:
        pass

    try:
        row = db.execute("""
            SELECT total_population, aging_rate
            FROM v2_external_population WHERE prefecture = '北海道' AND municipality = '札幌市'
        """).fetchone()
        if row:
            print(f"  北海道札幌市: 人口{row[0]:,}人, 高齢化率{row[1]:.1f}%")
    except Exception:
        pass

    try:
        row = db.execute("""
            SELECT inflow, outflow, net_migration, net_migration_rate
            FROM v2_external_migration WHERE prefecture = '東京都' AND municipality = '特別区部'
        """).fetchone()
        if row:
            print(f"  東京都特別区部: 転入{row[0]:,}人, 転出{row[1]:,}人, 社会増減{row[2]:+,}人 ({row[3]:+.1f}permille)")
    except Exception:
        pass

    try:
        row = db.execute("""
            SELECT total_foreign, foreign_rate
            FROM v2_external_foreign_residents WHERE prefecture = '東京都' AND municipality = '新宿区'
        """).fetchone()
        if row:
            print(f"  東京都新宿区: 外国人{row[0]:,}人, 比率{row[1]:.1f}%")
    except Exception:
        pass

    try:
        row = db.execute("""
            SELECT nighttime_pop, daytime_pop, day_night_ratio
            FROM v2_external_daytime_population WHERE prefecture = '東京都' AND municipality = '千代田区'
        """).fetchone()
        if row:
            print(f"  東京都千代田区: 夜間{row[0]:,}人, 昼間{row[1]:,}人, 昼夜比{row[2]:.1f}%")
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser(description="SSDSE-A CSV -> hellowork.db インポート")
    parser.add_argument("--csv", default=DEFAULT_CSV, help="SSDSE-A CSVファイルパス")
    parser.add_argument("--db", default=DEFAULT_DB, help="hellowork.dbパス")
    parser.add_argument("--daytime-xlsx", default=DEFAULT_DAYTIME_XLSX,
                        help="昼夜間人口Excelパス (e-Stat 国勢調査)")
    parser.add_argument("--census-age-xlsx", default=DEFAULT_CENSUS_AGE_XLSX,
                        help="国勢調査 年齢各歳別人口Excel (10歳階級按分用)")
    args = parser.parse_args()

    if not os.path.exists(args.csv):
        print(f"ERROR: CSV not found: {args.csv}")
        print("  SSDSE-A-2025.csv を以下からダウンロードしてください:")
        print("  https://www.nstac.go.jp/files/SSDSE-A-2025.csv")
        print(f"  配置先: {DEFAULT_CSV}")
        sys.exit(1)

    if not os.path.exists(args.db):
        print(f"ERROR: DB not found: {args.db}")
        sys.exit(1)

    print(f"CSV: {args.csv}")
    print(f"DB:  {args.db}")
    print()

    # CSV読み込み
    rows, ref_years = read_ssdse_csv(args.csv)
    if rows is None:
        sys.exit(1)

    print(f"  参照年: {ref_years}")
    print()

    # DB書き込み
    db = sqlite3.connect(args.db)
    db.execute("PRAGMA journal_mode=WAL")

    try:
        print("インポート開始...")

        # 1. 人口（他テーブルが分母として使う）
        pop_rows = import_population(db, rows, ref_years)
        db.commit()

        # 2. 社会動態
        import_migration(db, rows, ref_years, pop_rows)
        db.commit()

        # 3. 外国人住民
        import_foreign_residents(db, rows, ref_years, pop_rows)
        db.commit()

        # 4. 人口ピラミッド（9区分×男女、国勢調査按分）
        census_ratios = parse_census_age_excel(args.census_age_xlsx)
        import_population_pyramid(db, rows, census_ratios)
        db.commit()

        # 5. 昼夜間人口（e-Stat 国勢調査 従業地・通学地集計 令和2年）
        code_map = build_code_to_name_map(args.csv)
        import_daytime_population(db, args.daytime_xlsx, code_map)
        db.commit()

        # 検証
        verify(db)

        print("\nSSDSE-Aインポート完了")
        print("  ベンチマーク12軸の再計算: python compute_v2_external.py")

    except Exception as e:
        db.rollback()
        print(f"ERROR: {e}")
        raise
    finally:
        db.close()


if __name__ == "__main__":
    main()
