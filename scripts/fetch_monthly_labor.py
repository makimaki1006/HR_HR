"""
毎月勤労統計調査 地方調査 (都道府県別 賃金・労働時間) fetch スクリプト

出典: 厚生労働省「毎月勤労統計調査 地方調査」(全国・都道府県別統計表)
      https://www.mhlw.go.jp/toukei/list/30-1.html
      直DL: https://www.mhlw.go.jp/toukei/itiran/roudou/monthly/r{RR}/{RRMM}/xlsx/{RRMM}.xlsx
利用規約: 政府標準利用規約(第2.0版)準拠。出典明記のうえ複製・公衆送信・翻案可。
          (最終的なライセンス条項は WF12 調査結果に従って確認すること。caveats 参照)

処理概要:
  - MHLW サイトから月次 Excel を直接ダウンロード (e-Stat API 不使用のため appId 不要)
  - 直近 12 ヶ月分をバックフィル (404 はスキップして caveats に記録)
  - 表1 (規模5人以上) / 表2 (規模30人以上) の「調査産業計」都道府県別値を抽出
  - 表3 / 表4 は労働者数のみのため対象外
  - 注記: Excel の「全国」行は全国調査の結果であり、都道府県別地方調査の平均・合計ではない
          (Excel 注記 1 参照)。本スクリプトは prefecture='全国' 行として区別して出力。

出力: scripts/staging/monthly_labor.csv (UTF-8, ヘッダ付き)
  列: prefecture, year_month, industry, size_class,
      cash_earnings_total, scheduled_earnings, total_hours, overtime_hours

Excel 表構造 (表1/表2 共通, data_only):
  行2: "2025(令和7)年7月"        -> year_month の抽出元
  行4-5: ヘッダ + 単位
  行6: 全国
  行8-63: 47 都道府県 (空行区切りあり)
  列A=都道府県, C=総実労働時間, E=所定外労働時間,
  列G=現金給与総額, I=所定内給与
"""
import argparse
import csv
import io
import re
import sys
from datetime import date
from pathlib import Path

import openpyxl
import requests

sys.stdout.reconfigure(encoding="utf-8")

BASE_URL = "https://www.mhlw.go.jp/toukei/itiran/roudou/monthly"
OUTPUT_PATH = Path(__file__).parent / "staging" / "monthly_labor.csv"
INDUSTRY = "調査産業計"

# 表 -> (シート接尾辞, size_class)。表3/表4(労働者数のみ)は対象外
TABLES = [("T1", "5人以上"), ("T2", "30人以上")]

# 列マッピング (1-indexed)
COL_PREF = 1
COL_TOTAL_HOURS = 3        # 総実労働時間 (時間)
COL_OVERTIME_HOURS = 5     # 所定外労働時間 (時間)
COL_CASH_EARNINGS = 7      # 現金給与総額 (円)
COL_SCHEDULED_EARNINGS = 9  # 所定内給与 (円)

DATA_ROW_START = 6   # 全国
DATA_ROW_END = 63    # 沖縄

PREFS_47 = {
    "北海道", "青森", "岩手", "宮城", "秋田", "山形", "福島", "茨城", "栃木",
    "群馬", "埼玉", "千葉", "東京", "神奈川", "新潟", "富山", "石川", "福井",
    "山梨", "長野", "岐阜", "静岡", "愛知", "三重", "滋賀", "京都", "大阪",
    "兵庫", "奈良", "和歌山", "鳥取", "島根", "岡山", "広島", "山口", "徳島",
    "香川", "愛媛", "高知", "福岡", "佐賀", "長崎", "熊本", "大分", "宮崎",
    "鹿児島", "沖縄",
}


def reiwa_code(year: int, month: int) -> str:
    """西暦年月 -> 令和 YYMM 文字列 (令和1 = 2019)。"""
    r = year - 2018
    return f"{r:02d}{month:02d}"


def month_url(year: int, month: int) -> str:
    rr = f"{year - 2018:02d}"
    rrmm = reiwa_code(year, month)
    return f"{BASE_URL}/r{rr}/{rrmm}/xlsx/{rrmm}.xlsx"


def iter_months_back(end_year: int, end_month: int, n: int):
    """end から遡って n ヶ月分 (year, month) を新しい順に返す。"""
    y, m = end_year, end_month
    for _ in range(n):
        yield y, m
        m -= 1
        if m == 0:
            m = 12
            y -= 1


def detect_latest(probe_from: date, max_back: int = 18):
    """probe_from から遡って最初に 200 を返す月を最新とみなす。"""
    y, m = probe_from.year, probe_from.month
    for _ in range(max_back):
        url = month_url(y, m)
        try:
            r = requests.head(url, timeout=20, allow_redirects=True)
            if r.status_code == 200:
                return y, m
        except requests.RequestException:
            pass
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    return None


def clean_pref(v: str) -> str:
    return v.replace("　", "").replace(" ", "").strip()


def parse_year_month(ws) -> str:
    """行2 '2025(令和7)年7月' -> '2025-07'。"""
    raw = ws.cell(2, 1).value or ""
    mo = re.search(r"(\d{4}).*?年\s*(\d{1,2})\s*月", str(raw))
    if not mo:
        return ""
    return f"{int(mo.group(1)):04d}-{int(mo.group(2)):02d}"


def parse_sheet(ws, size_class: str, expected_ym: str, rows_out: list, warnings: list):
    ym = parse_year_month(ws)
    if ym and expected_ym and ym != expected_ym:
        warnings.append(f"year_month不一致: sheet={ym} expected={expected_ym}")
    if not ym:
        ym = expected_ym
    for r in range(DATA_ROW_START, DATA_ROW_END + 1):
        pref = ws.cell(r, COL_PREF).value
        if not pref:
            continue
        pref = clean_pref(str(pref))
        cash = ws.cell(r, COL_CASH_EARNINGS).value
        sched = ws.cell(r, COL_SCHEDULED_EARNINGS).value
        total_h = ws.cell(r, COL_TOTAL_HOURS).value
        over_h = ws.cell(r, COL_OVERTIME_HOURS).value
        if cash is None and total_h is None:
            continue
        rows_out.append({
            "prefecture": pref,
            "year_month": ym,
            "industry": INDUSTRY,
            "size_class": size_class,
            "cash_earnings_total": cash,
            "scheduled_earnings": sched,
            "total_hours": total_h,
            "overtime_hours": over_h,
        })


def validate(rows: list, warnings: list):
    """ドメイン検証。"""
    # 月×規模ごとに 47 都道府県が揃うか
    groups = {}
    for row in rows:
        key = (row["year_month"], row["size_class"])
        groups.setdefault(key, set()).add(row["prefecture"])
    full_month_found = False
    for key, prefs in sorted(groups.items()):
        present = prefs & PREFS_47
        if len(present) == 47:
            full_month_found = True
        else:
            missing = PREFS_47 - prefs
            warnings.append(f"{key}: 47都道府県未充足 present={len(present)} missing={sorted(missing)}")
    if not full_month_found:
        warnings.append("警告: 47都道府県が揃う月×規模が存在しない")

    # 値域チェック
    # 現金給与総額は 6月/12月 の賞与(特別給与)で大きく増えるため広めのレンジを許容。
    # 賞与を含まない所定内給与は安定レンジで別途チェックする。
    for row in rows:
        if row["prefecture"] == "全国":
            continue
        cash = row["cash_earnings_total"]
        if cash is not None and not (150000 <= cash <= 1_100_000):
            warnings.append(f"cash_earnings 常識レンジ外: {row['prefecture']} {row['year_month']} {row['size_class']} = {cash}")
        sched = row["scheduled_earnings"]
        if sched is not None and not (150000 <= sched <= 500000):
            warnings.append(f"scheduled_earnings 常識レンジ外: {row['prefecture']} {row['year_month']} {row['size_class']} = {sched}")
        th = row["total_hours"]
        if th is not None and not (100 <= th <= 200):
            warnings.append(f"total_hours 範囲外: {row['prefecture']} {row['year_month']} {row['size_class']} = {th}")
        oh = row["overtime_hours"]
        if oh is not None and not (0 <= oh <= 40):
            warnings.append(f"overtime_hours 範囲外: {row['prefecture']} {row['year_month']} {row['size_class']} = {oh}")
    return full_month_found


def main():
    ap = argparse.ArgumentParser(description="毎月勤労統計 地方調査 fetch")
    ap.add_argument("--end", help="最新月 YYYY-MM (省略時は自動検出)")
    ap.add_argument("--months", type=int, default=12, help="バックフィル月数 (default 12)")
    args = ap.parse_args()

    if args.end:
        ey, em = map(int, args.end.split("-"))
    else:
        latest = detect_latest(date.today())
        if not latest:
            print("最新月を検出できませんでした", file=sys.stderr)
            sys.exit(1)
        ey, em = latest
        print(f"自動検出した最新月: {ey}-{em:02d}")

    rows = []
    warnings = []
    fetched, skipped = [], []

    for y, m in iter_months_back(ey, em, args.months):
        ym = f"{y:04d}-{m:02d}"
        url = month_url(y, m)
        try:
            resp = requests.get(url, timeout=60)
        except requests.RequestException as e:
            skipped.append(f"{ym} (取得エラー: {e})")
            continue
        if resp.status_code != 200:
            skipped.append(f"{ym} (HTTP {resp.status_code})")
            continue
        try:
            wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)
        except Exception as e:
            skipped.append(f"{ym} (Excel読込エラー: {e})")
            continue
        for suffix, size_class in TABLES:
            sheet = f"{reiwa_code(y, m)}{suffix}"
            if sheet not in wb.sheetnames:
                warnings.append(f"{ym}: シート {sheet} が存在しない")
                continue
            parse_sheet(wb[sheet], size_class, ym, rows, warnings)
        fetched.append(ym)
        print(f"取得: {ym} ({url})")

    full_ok = validate(rows, warnings)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    fields = ["prefecture", "year_month", "industry", "size_class",
              "cash_earnings_total", "scheduled_earnings", "total_hours", "overtime_hours"]
    with open(OUTPUT_PATH, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)

    print("\n===== サマリ =====")
    print(f"出力: {OUTPUT_PATH}")
    print(f"行数: {len(rows)}")
    print(f"取得月: {len(fetched)} {fetched}")
    print(f"スキップ月: {len(skipped)} {skipped}")
    print(f"47都道府県充足月あり: {full_ok}")
    print(f"警告 ({len(warnings)}):")
    for wn in warnings[:30]:
        print("  -", wn)
    print("\n先頭3行:")
    for row in rows[:3]:
        print("  ", row)


if __name__ == "__main__":
    main()
